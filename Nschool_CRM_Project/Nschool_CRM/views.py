
from decimal import Decimal
from io import BytesIO
import json, requests
from django.core.files.uploadedfile import InMemoryUploadedFile
from django.apps import apps

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth import authenticate
from django.contrib.auth import authenticate, login , logout as auth_logout
from django.template import RequestContext
from yaml import Loader
from .models import NewUser, Course
from django.contrib.auth.decorators import login_required

from django.core.cache import cache
from django.core.paginator import Paginator

from rest_framework import generics
from rest_framework.permissions import IsAuthenticated, AllowAny
from .serializer import NewUserSerializer

from rest_framework.authtoken.models import Token
from django.http import JsonResponse
from django.views.decorators.csrf import csrf_protect
from django.views.decorators.http import require_GET

from django.core.exceptions import ObjectDoesNotExist
from rest_framework.decorators import api_view, permission_classes
from rest_framework.response import Response
from rest_framework import status

from .serializer import *
from django.forms.models import model_to_dict

import csv
from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

from PIL import Image as PILImage
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter

import openpyxl
from django.http import Http404
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Font, Alignment, Border, Side
from PIL import Image as PILImage
from io import BytesIO
import re
from . import renderers

from .utils import calculate_payment_totals, encrypt_password, decrypt_password
from django.views.generic import TemplateView, ListView
from django.db.models import Q

from django.contrib import messages

from django.core.exceptions import ValidationError

from django.conf import settings

EMI_MODELS = getattr(settings, 'EMI_MODELS', {})

@csrf_protect
def admin_login(request):
    if request.method == 'POST':
        username = request.POST.get('username')
        email = request.POST.get('email')
        password = request.POST.get('password')
        
        print(email, password)
        
        url = 'http://127.0.0.1:8000/api/login/'
        # data = {'username': username, 'password': password}
        data = {'username_or_email': email, 'password': password, 'username': username}
        csrf_token = request.COOKIES.get('csrftoken')
        
        try:
            headers = {
                'X-CSRFToken': csrf_token
            }
            response = requests.post(url, data=data, headers=headers)
            # response.raise_for_status()  # Raise an HTTPError for bad responses
            response_data = response.json()
    
        except requests.exceptions.RequestException as e:
            context = {
                'errors': f'Error occurred: {e}'
            }
            return render(request, 'admin_login.html', context)
        
        if response.status_code == 200:
            user = authenticate(request, email=email, password=password)
            if user is not None:
                login(request, user)
                return redirect('dashboard')
            else:
                context = {'error': 'Invalid credentials'}
                return render(request, 'admin_login.html', context)
        else:
            context = {
                'error': response_data.get('user_error', response_data.get('pass_error', response_data.get('error', 'Invalid credentials')))
            }
            return render(request, 'admin_login.html', context)
    
    return render(request, 'admin_login.html')

def logout(request):
    token = Token.objects.get()
    api_url = 'http://127.0.0.1:8000/api/logout/'
    headers = {
        'Authorization': f'Token {token.key}'
    }
    
    try:
        response = requests.post(api_url, headers=headers)
        response.raise_for_status()
    except requests.exceptions.RequestException as e:
        print(f'Error during API logout: {e}')
        return redirect('dashboard')  # Redirect to dashboard or show an error message

    # Remove the token locally after successful API logout
    token.delete()

    # Log out the user and clear the session
    auth_logout(request)
    request.session.flush()

    return redirect('admin_login')

def dashboard_view(request):
    datapoints = [
        { "x": 10, "y": 171 },
        { "x": 20, "y": 155},
        { "x": 30, "y": 150 },
        { "x": 40, "y": 165 },
        { "x": 50, "y": 195 },
        { "x": 60, "y": 168 },
        { "x": 70, "y": 128 },
        { "x": 80, "y": 134 },
        { "x": 90, "y": 114}
    ]
 
    datapoints2 = [
        { "x": 10, "y": 71 },
        { "x": 20, "y": 55},
        { "x": 30, "y": 50 },
        { "x": 40, "y": 65 },
        { "x": 50, "y": 95 },
        { "x": 60, "y": 68 },
        { "x": 70, "y": 28 },
        { "x": 80, "y": 34 },
        { "x": 90, "y": 14 }
    ]
    
    return render(request, 'dashboard.html',  { "datapoints" : json.dumps(datapoints), "datapoints2": json.dumps(datapoints2) })

def user_module_insert_view(request):
    if request.method == 'POST':
        # Extract data from the form
        username = request.POST.get("username", "").strip()
        email = request.POST.get("email", "").strip()
        contact = request.POST.get("contact", "").strip()
        designation = request.POST.get("designation", "").strip()
        password = request.POST.get("password", "").strip()
        cpassword = request.POST.get("cpassword", "").strip()
        
        # Extract permission fields
        enquiry = "Enquiry" in request.POST
        enrollment = "Enrollment" in request.POST
        payment = "Payment" in request.POST
        attendance = "Attendance" in request.POST
        staff = "Staff" in request.POST
        placement = "Placement" in request.POST
        report = "Report" in request.POST
        
        # Validate password
        if password != cpassword:
            context = {
                'error': 'Passwords do not match'
            }
            return render(request, 'new_user.html', context)


        # Encrypt the password
        encrypted_password = encrypt_password(password)
        
        
        # Prepare data for the API request
        user_data = {
            'username': username,
            'email': email.lower(),
            'contact': contact,
            'designation': designation,
            'password': encrypted_password.decode(),
            'enquiry': "Enquiry" in request.POST,
            'enrollment': "Enrollment" in request.POST,
            'payment': "Payment" in request.POST,
            'attendance': "Attendance" in request.POST,
            'staff': "Staff" in request.POST,
            'placement': "Placement" in request.POST,
            'report': "Report" in request.POST,
        }

        # Get the token
        try:
            token = Token.objects.first()
        except Token.DoesNotExist:
            context = {
                'error': 'Authentication token not found'
                **user_data,
            }
            return render(request, 'new_user.html', context)
        
        api_url = 'http://127.0.0.1:8000/api/newuser/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }

        try:
            response = requests.post(api_url, json=user_data, headers=headers)
            response_data = response.json()
            print(f"Response Data : {response_data}")
        except requests.exceptions.HTTPError as http_err:
            # Handle specific HTTP errors
            context = {
                'error': f'HTTP error occurred: {http_err}',
                **user_data,
                'response_data': response.json()
            }
            return render(request, 'new_user.html', context)
        except requests.exceptions.RequestException as req_err:
            # Handle general request exceptions
            print(f'Error during API create New User: {req_err}')
            context = {
                'error': 'An error occurred while creating the new user.'
            }
            return render(request, 'new_user.html', context)        
        
        if response.status_code == 201:
            context =  {
                "message": "Created Successfully"
            }
            return render(request, 'new_user.html', context)
        else:
            error_messages = {
                'username': response_data.get('username', ''),
                'email': response_data.get('email', ''),
                'contact': response_data.get('contact', ''),
                'designation': response_data.get('designation', ''),
                'password': response_data.get('password', ''),
            }
            context = {
                'username': username,
                'email': email if not error_messages.get('email') else '',  # Clear if there's an email error
                'contact': contact if not error_messages.get('contact') else '',  # Clear if there's a contact error
                'designation': designation,
                'enquiry': enquiry,
                'enrollment': enrollment,
                'payment': payment, 
                'attendance': attendance,
                'staff': staff,
                'placement': placement,
                'report': report,
                'password': password,
                'cpassword': cpassword,
                'error_messages': error_messages,
                
            }
            return render(request, 'new_user.html', context)
        
    return render(request, 'new_user.html')

def manage_user_view(request):
    # Fetch the token
    try:
        token = Token.objects.first()  # Assuming you only have one token and it's safe to get the first one
    except Token.DoesNotExist:
        context = {
            'error': 'Authentication token not found'
        }
        return render(request, 'manage_user.html', context)
    
    api_url = 'http://127.0.0.1:8000/api/newuser/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        response_data = response.json()
        
        # Decrypt the passwords
        for user_data in response_data:
            encrypted_password = user_data.get('password')
            if encrypted_password:
                try:
                    user_data['password'] = decrypt_password(encrypted_password)
                except Exception as e:
                    user_data['password'] = 'Error decrypting password' 
        
    except requests.exceptions.RequestException as err:
        # Catch any request-related exceptions
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_user.html', context)

    # Get the per_page value from the request, default to 10 if not provided
    per_page = request.GET.get('per_page', '10')

    # Apply pagination
    paginator = Paginator(response_data, per_page)  # Use response_data for pagination
    
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
    }
    return render(request, 'manage_user.html', context)

def delete_user_view(request, id):
    user_id = NewUser.objects.get(id=id)
    
    print(user_id.pk)
    
    if not user_id:
        context = {'error': 'User ID not provided'}
        print("user id not provided")
        return render(request, 'manage_user.html', context)
    
    try:
        token = Token.objects.get()  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_user.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/newuser/{user_id.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()
        print(response)
    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_user.html', context)
    
    if response.status_code == 204:
        return redirect('manage-user')
    
    else:
        response_data = response.json()
        context = {
            'detail': response_data.get('detail', 'An error occurred while deleting the user'),
        }
        return render(request, 'manage_user.html', context)

def delete_all_users_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_ids = data.get('user_ids', [])
            if user_ids:
                NewUser.objects.filter(id__in=user_ids).delete()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'No users selected for deletion'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})
    

def update_user_view(request, id):
    try:
        user = NewUser.objects.get(id=id)
    except NewUser.DoesNotExist:
        context = {'error': 'User not found'}
        return render(request, 'manage_user.html', context)

    if request.method == 'POST':
        print("User before update:", user)
        
        try:
            token = Token.objects.first()  # Get the first token for simplicity
            if not token:
                raise Token.DoesNotExist
        except Token.DoesNotExist:
            context = {'error': 'Authentication token not found'}
            return render(request, 'manage_user.html', context)
        
        api_url = f'http://127.0.0.1:8000/api/update_newuser/{user.pk}/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }
        
        user_data = {
            'username': request.POST.get('username', user.username),
            'email': request.POST.get('email', user.email),
            'contact': request.POST.get('contact', user.contact),
            'designation': request.POST.get('designation', user.designation),
            'enquiry': 'Enquiry' in request.POST,
            'enrollment': 'Enrollment' in request.POST,
            'payment': 'Payment' in request.POST,
            'attendance': 'Attendance' in request.POST,
            'staff': 'Staff' in request.POST,
            'placement': 'Placement' in request.POST,
            'report': 'Report' in request.POST,
        }
        
        print(user_data['enquiry'])
        
        print("User Data being sent:", user_data)

        try:
            response = requests.patch(api_url, data=json.dumps(user_data), headers=headers)
            print("API Response Status Code:", response.status_code)
            response.raise_for_status()
            response_data = response.json()
            print("API Response Data:", response_data)
        except requests.exceptions.RequestException as err:
            context = {
                'error': f'Request error occurred: {err}',
                'response_data': response.json() if response.content else {}
            }
            return render(request, 'manage_user.html', context)
        
        if response.status_code in [200, 204]:  # 204 No Content is also a valid response for updates
            print("Update successful")
            return redirect('manage-user')
        else:
            context = {
                'error': 'Failed to update user information',
                'username': response_data.get('username', ''),
                'email': response_data.get('email', ''),
                'contact': response_data.get('contact', ''),
                'designation': response_data.get('designation', ''),
            }
            return render(request, 'update_user.html', context)
        
    return render(request, 'update_user.html', {"user": user})
    
# create APIS

# user login api

import logging

logger = logging.getLogger(__name__)

@api_view(['POST'])
@permission_classes([AllowAny])
def user_login(request):
    serializer = LoginSerializer(data=request.data)
    
    print("serializer : ", serializer)

    if serializer.is_valid():
        username_or_email = serializer.validated_data['username_or_email']
        password = serializer.validated_data['password']

        print("email : ", username_or_email)
        print("password : ", password)

        user = authenticate(email=username_or_email, password=password)
        
        print(user)
        
        print(" Login user : ", user)
        
        if user:
            if not isinstance(user, (NewUser)):
                print("not a isinstance")
                return Response({'error': 'Invalid user type'}, status=status.HTTP_401_UNAUTHORIZED)

            try:
                token, created = Token.objects.get_or_create(user=user)
                print("Token : ", token)
            except Exception as e:
                return Response({'error': str(e)}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)

            return Response({'token': token.key}, status=status.HTTP_200_OK)
        else:
            return Response({'error': 'Invalid credentials'}, status=status.HTTP_401_UNAUTHORIZED)

    return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

# user logout api view
@api_view(["POST",])
@permission_classes([IsAuthenticated])
def user_logout(request):
    if request.method == "POST":
        request.user.auth_token.delete()
        return Response({"Message": "You are logged out"}, status=status.HTTP_200_OK)

# New user APi view

class NewUserListCreateView(generics.ListCreateAPIView):
    queryset = NewUser.objects.all().order_by('-id')
    serializer_class = NewUserSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'Message': 'No users found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

    def create(self, request, *args, **kwargs):
        email = request.data.get('email')
        contact = request.data.get('contact')
        
        errors = {}
        
        if NewUser.objects.filter(email=email).exists():
            errors['email'] = 'This Email ID is Already Exist.'

        if NewUser.objects.filter(contact=contact).exists():
            errors['contact'] = 'This Contact Number is Already Exist.'
            
        # If there are any errors, return them
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        serializer = self.get_serializer(data=request.data)
        try:
            serializer.is_valid(raise_exception=True)
        except serializers.ValidationError as e:
            custom_errors = {}
            for field, messages in e.detail.items():
                if field == 'name':
                    custom_errors[field] = 'Please provide a valid name.'
                elif field == 'email':
                    custom_errors[field] = 'Email is required and must be unique.'
                elif field == 'contact':
                    custom_errors[field] = 'Contact number must be valid and unique.'
                else:
                    custom_errors[field] = messages[0]  # Fallback for other fields

            return Response(custom_errors, status=status.HTTP_400_BAD_REQUEST)
        
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)
    
class NewUserDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = NewUser.objects.all()
    serializer_class = NewUserSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class NewUserUpdateView(generics.RetrieveUpdateAPIView):
    queryset = NewUser.objects.all()
    serializer_class = NewUserUpdateSerializer
    permission_classes = [IsAuthenticated]
    partial = True
    
    
class NewUserDeleteView(generics.DestroyAPIView):
    queryset = NewUser.objects.all()
    serializer_class = NewUserSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

# New Course APi view

class CourseListCreateView(generics.ListCreateAPIView):
    queryset = Course.objects.all().order_by('-created_at')
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'Message': 'No users found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class CourseDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class CourseUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]
    partial = True
    
    
class CourseDeleteView(generics.DestroyAPIView):
    queryset = Course.objects.all()
    serializer_class = CourseSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

# New Enquiry Mode APi view

class EnquiryModeListCreateView(generics.ListCreateAPIView):
    queryset = Enquiry_Mode.objects.all().order_by('-id')
    serializer_class = EnquiryModeSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'Message': 'No users found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    
    def create(self, request, *args, **kwargs):
        print(request.data)  # Log incoming data for debugging
        email_id = request.data.get('email_id')
        contact_no = request.data.get('contact_no')
        
        errors = {}
        
        if NewUser.objects.filter(email_id=email_id).exists():
            errors['email'] = 'This Email ID already exists.'

        if NewUser.objects.filter(contact_no=contact_no).exists():
            errors['contact'] = 'This Contact Number already exists.'
        
        # If there are any errors, return them
        if errors:
            return Response(errors, status=status.HTTP_400_BAD_REQUEST)

        # Create and validate the serializer with the request data
        serializer = self.get_serializer(data=request.data)
        
        # Call is_valid to trigger validation including your date validation
        serializer.is_valid(raise_exception=True)

        # If valid, perform the creation
        self.perform_create(serializer)
        return Response(serializer.data, status=status.HTTP_201_CREATED)


class EnquiryModeDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Enquiry_Mode.objects.all()
    serializer_class = EnquiryModeSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EnquiryModeUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Enquiry_Mode.objects.all()
    serializer_class = EnquiryModeSerializer
    permission_classes = [IsAuthenticated]
    partial = True
    
    
class EnquiryModeDeleteView(generics.DestroyAPIView):
    queryset = Enquiry_Mode.objects.all()
    serializer_class = EnquiryModeSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

# New Enquiry APi view

class EnquiryListCreateView(generics.ListCreateAPIView):
    queryset = Enquiry.objects.all().order_by('-id')
    serializer_class = EnquirySerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'Message': 'No users found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class EnquiryDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Enquiry.objects.all()
    serializer_class = EnquirySerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EnquiryUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Enquiry.objects.all()
    serializer_class = EnquirySerializer
    permission_classes = [IsAuthenticated]
    partial = True    
    
class EnquiryDeleteView(generics.DestroyAPIView):
    queryset = Enquiry.objects.all()
    serializer_class = EnquirySerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

# Api for notes
class NotesDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Notes.objects.all()
    serializer_class = NotesSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'
    
class NotesUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Notes.objects.all()
    serializer_class = NotesSerializer
    permission_classes = [IsAuthenticated]
    partial = True
    
class NotesDeleteView(generics.DestroyAPIView):
    queryset = Notes.objects.all()
    serializer_class = NotesSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})
    
# Enrollment Api

class EnrollmentListCreateView(generics.ListCreateAPIView):
    queryset = Enrollment.objects.all().order_by('-id')
    serializer_class = EnrollmentSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'Message': 'No Enrollment found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class EnrollmentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EnrollmentUpdateView(generics.RetrieveUpdateAPIView):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    permission_classes = [IsAuthenticated]
    partial = True    
    
class EnrollmentDeleteView(generics.DestroyAPIView):
    queryset = Enrollment.objects.all()
    serializer_class = EnrollmentSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

# payment info API

class PaymentInfoListCreateView(generics.ListCreateAPIView):
    queryset = PaymentInfo.objects.prefetch_related('single_payment', 'emi_1_payments').all().order_by('-id')
    serializer_class = PaymentInfoSerializer
    permission_classes = [IsAuthenticated]

    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        logger.debug("Queryset: %s", queryset)
        if not queryset.exists():
            return Response({'Message': 'No Payment Records found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        logger.debug("Serialized Data: %s", serializer.data)
        return Response(serializer.data, status=status.HTTP_200_OK)

class PaymentInfoDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = PaymentInfo.objects.prefetch_related('installments', 'single_payment').all()  # If you still need the relations in certain cases
    serializer_class = PaymentInfoSerializer  # Use the serializer that excludes related fields by default
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class PaymentInfoUpdateView(generics.RetrieveUpdateAPIView):
    queryset = PaymentInfo.objects.all()
    serializer_class = PaymentInfoSerializer
    permission_classes = [IsAuthenticated]
    partial = True
    
class PaymentInfoDeleteView(generics.DestroyAPIView):
    queryset = PaymentInfo.objects.all()
    serializer_class = PaymentInfoSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

# Installment API

class SinglePaymentListCreateView(generics.ListCreateAPIView):
    queryset = SinglePayment.objects.all().order_by('-id')
    serializer_class = SinglePaymentSerializer
    permission_classes = [IsAuthenticated]
    
    def list(self, request, *args, **kwargs):
        queryset = self.get_queryset()
        if not queryset.exists():
            return Response({'Message': 'No Payment Records found'}, status=status.HTTP_404_NOT_FOUND)
        
        serializer = self.get_serializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)

class SinglePaymentDetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = SinglePayment.objects.all()
    serializer_class = SinglePaymentSerializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class SinglePaymentUpdateView(generics.RetrieveUpdateAPIView):
    queryset = SinglePayment.objects.all()
    serializer_class = SinglePaymentSerializer
    permission_classes = [IsAuthenticated]
    partial = True
    
class SinglePaymentDeleteView(generics.DestroyAPIView):
    queryset = SinglePayment.objects.all()
    serializer_class = SinglePaymentSerializer
    permission_classes = [IsAuthenticated]

    def destroy(self, request, *args, **kwargs):
        instance = self.get_object()
        instance.delete()
        return Response({'Message': 'Successfully deleted'})

@csrf_exempt
def export_user_csv(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="course_list_csv.csv"'

        writer = csv.writer(response)

        # Write the header row
        writer.writerow(['Name', 'Email', 'Contact', 'Designation', 'Permission'])

        # Fetch selected courses based on IDs
        selected_courses = NewUser.objects.filter(id__in=ids)

        for user in selected_courses:
            # Remove country code from contact number
            contact_number = str(user.contact)
            # contact_number = re.sub(r'^\+\d{1,2}', '', contact_number)
            
            permission = []
                
            if user.enquiry == True:
                permission.append('Enquiry')
            
            if user.enrollment == True:
                permission.append('Enrollment')
            
            if user.enrollment == True:
                permission.append('Enrollment')
            
            if user.attendance == True:
                permission.append('Attendance')
                
            if user.placement == True:
                permission.append('Permission')
            
            if user.staff == True:
                permission.append('Staff')
            
            if user.report == True:
                permission.append('Report')
                    
            # Write the data row
            writer.writerow([user.username, user.email, contact_number, user.designation, ', '.join(permission)])

        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX



from django.http import HttpResponse
from django.views.decorators.csrf import csrf_exempt

@csrf_exempt
def export_user_excel(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Fetch selected courses based on IDs
        selected_courses = NewUser.objects.filter(id__in=ids)
        
        if not selected_courses:
            return JsonResponse({'error': 'No users available.'}, status=404)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define header row with font style and alignment
        header_row = ['Name', 'Email', 'Contact', 'Designation', 'Permission']
        ws.append(header_row)

        for idx, user in enumerate(selected_courses, start=2):
                
            permission = []
                
            if user.enquiry == True:
                permission.append('Enquiry')
            
            if user.enrollment == True:
                permission.append('Enrollment')
            
            if user.attendance == True:
                permission.append('Attendance')
                
            if user.placement == True:
                permission.append('Permission')
            
            if user.staff == True:
                permission.append('Staff')
            
            if user.report == True:
                permission.append('Report')
            
            ws.append([user.username, user.email, user.contact, user.designation, ', '.join(permission)])

        # Create an in-memory file-like object to save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response with Excel content type and attachment header
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
        
        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

from django.http import HttpResponse
from xhtml2pdf import pisa
from django.views.decorators.http import require_POST

@csrf_protect
@require_POST
def export_user_pdf(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        selected_users = NewUser.objects.filter(id__in=ids)
        
        if not selected_users:
            return JsonResponse({'error': 'No users available.'}, status=404)
        
        content_list = []
        for user in selected_users:
            # contact_number = re.sub(r'^\+\d{1,2}', '', contact_number)
                
            permission = []
                
            if user.enquiry == True:
                permission.append('Enquiry')
            
            if user.enrollment == True:
                permission.append('Enrollment')
            
            if user.attendance == True:
                permission.append('Attendance')
                
            if user.placement == True:
                permission.append('Permission')
            
            if user.staff == True:
                permission.append('Staff')
            
            if user.report == True:
                permission.append('Report')
                
            content_list.append({
                'username': user.username, 
                'email': user.email,
                'contact': user.contact,
                'designation': user.designation,
                'permission': ', '.join(permission),
            })
        
        content = {'user_list': content_list}
        return renderers.render_to_pdf('user_data_list.html', content)
    
    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX


class SearchResultsView(ListView):
    model = NewUser
    template_name = 'search_result.html'

    def get_queryset(self):
        query = self.request.GET.get("q")
        query_filter = Q()

        if query:
            query_lower = query.lower()
            
            boolean_fields = {
                'enquiry': 'enquiry',
                'enrollment': 'enrollment',
                'attendance': 'attendance',
                'staff': 'staff',
                'placement': 'placement',
                'report': 'report'
            }
            
            match_found = False
            for keyword, field in boolean_fields.items():
                if query_lower in keyword:
                    query_filter |= Q(**{f"{field}": True})
                    match_found = True
                    break

            if not match_found:
                # Search across CharField, EmailField, and PhoneNumberField
                fields = [f.name for f in NewUser._meta.fields if isinstance(f, (models.CharField, models.EmailField))]
                for field in fields:
                    query_filter |= Q(**{f"{field}__icontains": query})
        else:
            query_filter = Q(pk__isnull=True)

        object_list = NewUser.objects.filter(query_filter)
        return object_list


# course module

def generate_new_course_no():
    existing_courses = Course.objects.values_list('S_no', flat=True)
    
    if not existing_courses:
        return 1  # Start with 1 if there are no existing courses

    # Create a sorted list of existing S_no values
    existing_s_no = sorted(existing_courses)

    # Find the first available S_no
    for i in range(1, len(existing_s_no) + 2):
        if i not in existing_s_no:
            return i  # Return the first missing number
    
    return existing_s_no[-1] + 1  # Otherwise, increment the highest

def add_course_view(request):
    if request.method == 'POST':
        s_no = generate_new_course_no()
        # Extract data from the form
        course_name = request.POST.get("course", "").strip()
        
        # Prepare data for the API request
        
        user_data = {
            'course_name': course_name,
            'S_no' : s_no,
        }

        # Get the token
        try:
            token = Token.objects.first()
        except Token.DoesNotExist:
            context = {
                'error': 'Authentication token not found'
            }
            return render(request, 'add_course.html', context)
        
        api_url = 'http://127.0.0.1:8000/api/course/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }

        try:
            response = requests.post(api_url, json=user_data, headers=headers)
            response_data = response.json()
            
            print(response_data)
            
        except requests.exceptions.HTTPError as http_err:
            # Handle specific HTTP errors
            context = {
                'error': f'HTTP error occurred: {http_err}',
                'response_data': response.json()
            }
            return render(request, 'add_course.html', context)
        except requests.exceptions.RequestException as req_err:
            # Handle general request exceptions
            print(f'Error during API create Course: {req_err}')
            context = {
                'error': 'An error occurred while creating course.'
            }
            return render(request, 'add_course.html', context)        
        
        if response.status_code == 201:
            context =  {
                "message": "Created Successfully"
            }
            return render(request, 'add_course.html', context)
        else:
            context = {
                'course': response_data.get('course_name', ''),
            }
            return render(request, 'add_course.html', context)
        
    return render(request, 'add_course.html')

def manage_course_view(request):
    # Fetch the token
    try:
        token = Token.objects.first()  # Assuming you only have one token and it's safe to get the first one
    except Token.DoesNotExist:
        context = {
            'error': 'Authentication token not found'
        }
        return render(request, 'manage_user.html', context)
    
    api_url = 'http://127.0.0.1:8000/api/course/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        response_data = response.json()
        
    except requests.exceptions.RequestException as err:
        # Catch any request-related exceptions
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_course.html', context)

    # Get the per_page value from the request, default to 10 if not provided
    per_page = request.GET.get('per_page', '10')

    # Apply pagination
    paginator = Paginator(response_data, per_page)  # Use response_data for pagination
    
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
    }
    return render(request, 'manage_course.html', context)

from django.db.models import F

def delete_course_view(request, id):
    user_course = Course.objects.get(id=id)
    s_no_to_delete = user_course.S_no  # Get the S_no of the course to delete
    
    print(user_course.pk)
    
    if not user_course:
        context = {'error': 'User ID not provided'}
        print("user id not provided")
        return render(request, 'manage_course.html', context)
    
    try:
        token = Token.objects.get()  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_course.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/course/{user_course.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()
        
        # Adjust S_no for all courses with higher S_no
        Course.objects.filter(S_no__gt=s_no_to_delete).update(S_no=F('S_no') - 1)

    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_course.html', context)
    
    if response.status_code == 204:
        return redirect('manage-course')
    
    else:
        response_data = response.json()
        context = {
            'detail': response_data.get('detail', 'An error occurred while deleting the user'),
        }
        return render(request, 'manage_course.html', context)
    
def delete_all_course_view(request):
    if request.method == 'POST':
        print("Welcome")
        try:
            data = json.loads(request.body)
            
            print("Data : ", data)
            
            user_ids = data.get('user_ids', [])
            
            if user_ids:
                Course.objects.filter(id__in=user_ids).delete()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'No users selected for deletion'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

def update_course_view(request, id):
    try:
        user = Course.objects.get(id=id)
    except Course.DoesNotExist:
        context = {'error': 'Course not found'}
        return render(request, 'manage_course.html', context)

    if request.method == 'POST':
        
        try:
            token = Token.objects.first()  # Get the first token for simplicity
            if not token:
                raise Token.DoesNotExist
        except Token.DoesNotExist:
            context = {'error': 'Authentication token not found'}
            return render(request, 'manage_user.html', context)
        
        api_url = f'http://127.0.0.1:8000/api/update_course/{user.pk}/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }
        
        user_data = {
            'course_name': request.POST.get('course', user.course_name),
        }

        try:
            response = requests.patch(api_url, data=json.dumps(user_data), headers=headers)
            print("API Response Status Code:", response.status_code)
            response.raise_for_status()
            response_data = response.json()
            print("API Response Data:", response_data)
        except requests.exceptions.RequestException as err:
            context = {
                'error': f'Request error occurred: {err}',
                'response_data': response.json() if response.content else {}
            }
            return render(request, 'manage_course.html', context)
        
        if response.status_code in [200, 204]:  # 204 No Content is also a valid response for updates
            messages.error(request, "Update Successfully.")
            return redirect('manage-course')
        else:
            context = {
                'error': 'Failed to update user information',
                'course_name': response_data.get('course_name', ''),
            }
            return render(request, 'update_course.html', context)
        
    return render(request, 'update_course.html', {"course": user})

# csv file formate for course
@csrf_exempt
def export_course_csv(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="course_list_csv.csv"'

        writer = csv.writer(response)

        # Write the header row
        writer.writerow(['Course Name'])

        # Fetch selected courses based on IDs
        selected_courses = Course.objects.filter(id__in=ids)

        for user in selected_courses:        
            # Write the data row
            writer.writerow([user.course_name])

        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

# Excel file format for course
@csrf_exempt
def export_course_excel(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Fetch selected courses based on IDs
        selected_courses = Course.objects.filter(id__in=ids)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define header row with font style and alignment
        header_row = ['Course Name']
        ws.append(header_row)

        for idx, user in enumerate(selected_courses, start=2):
            ws.append([user.course_name])

        # Create an in-memory file-like object to save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response with Excel content type and attachment header
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
        
        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

@csrf_protect
@require_POST
def export_course_pdf(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        selected_users = Course.objects.filter(id__in=ids)
        
        if not selected_users:
            return JsonResponse({'error': 'No users available.'}, status=404)
        
        content_list = []
        for user in selected_users:    
            content_list.append({
                'course_name': user.course_name, 
            })
        
        content = {'course_list': content_list}
        return renderers.render_to_pdf('course_data_list.html', content)
    
    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

class SearchCourseResultsView(ListView):
    model = Course
    template_name = 'search_course_result.html'

    def get_queryset(self):
        query = self.request.GET.get("q")
        
        object_list = Course.objects.filter(
            Q(course_name__icontains = query)
        )
        
        return object_list

def generate_new_enquiry_no():
    try:
        last_enquiry = Enquiry.objects.latest('id')
        last_enquiry_no = last_enquiry.enquiry_no
    except Enquiry.DoesNotExist:
        last_enquiry_no = "EWT-0000" 

    numeric_part = int(last_enquiry_no.split('-')[1])
    incremented_numeric_part = numeric_part + 1
    return f"EWT-{incremented_numeric_part:04d}"

def generate_new_registration_no():
    try:
        last_enrollment = Enrollment.objects.latest('id')
        last_registration_no = last_enrollment.registration_no
    except Enrollment.DoesNotExist:
        last_registration_no = "EWRNO-0000" 

    numeric_part = int(last_registration_no.split('-')[1])
    incremented_numeric_part = numeric_part + 1
    return f"EWRNO-{incremented_numeric_part:04d}"

def add_attribute_view(request):
    if request.method == 'POST':
        # Extract data from the form
        mode_of_enquiry = request.POST.get("mode_of_enquiry", "").strip()
        
        # Prepare data for the API request
        user_data = {
            'mode_of_enquiry': mode_of_enquiry.lower(),
        }

        # Get the token
        try:
            token = Token.objects.get(user=request.user)
        except Token.DoesNotExist:
            context = {
                'error': 'Authentication token not found'
            }
            return render(request, 'add_attribute.html', context)
        
        api_url = 'http://127.0.0.1:8000/api/enquiry_mode/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }

        try:
            response = requests.post(api_url, json=user_data, headers=headers)
            response_data = response.json()
            
            print(response_data)
            
        except requests.exceptions.HTTPError as http_err:
            # Handle specific HTTP errors
            context = {
                'error': f'HTTP error occurred: {http_err}',
                'response_data': response.json()
            }
            return render(request, 'add_attribute.html', context)
        except requests.exceptions.RequestException as req_err:
            # Handle general request exceptions
            print(f'Error during API create Course: {req_err}')
            context = {
                'error': 'An error occurred while creating course.'
            }
            return render(request, 'add_attribute.html', context)        
        
        if response.status_code == 201:
            context =  {
                "message": "New Enquiry Mode Created Successfully"
            }
            return render(request, 'add_attribute.html', context)
        else:
            context = {
                'course': response_data.get('mode_of_enquiry', ''),
            }
            return render(request, 'add_attribute.html', context)
    return render(request, 'add_attribute.html')

def manage_attribute_view(request):
    # Fetch the token
    try:
        token = Token.objects.get(user=request.user)  # Assuming you only have one token and it's safe to get the first one
    except Token.DoesNotExist:
        context = {
            'error': 'Authentication token not found'
        }
        return render(request, 'manage_attributes.html', context)
    
    api_url = 'http://127.0.0.1:8000/api/enquiry_mode/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        response_data = response.json()
        
        print(response_data)
        
    except requests.exceptions.RequestException as err:
        # Catch any request-related exceptions
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_attributes.html', context)

    # Get the per_page value from the request, default to 10 if not provided
    per_page = request.GET.get('per_page', '10')

    # Apply pagination
    paginator = Paginator(response_data, per_page)  # Use response_data for pagination
    
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
    }
    return render(request, 'manage_attributes.html', context)

def delete_attribute_view(request, id):
    user_id = Enquiry_Mode.objects.get(id=id)
    
    print(user_id)
    
    if not user_id:
        context = {'error': 'Attribute ID not provided'}
        return render(request, 'manage_attributes.html', context)
    
    try:
        token = Token.objects.get(user=request.user)  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_attributes.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/enquiry_mode/{user_id.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()

    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_attributes.html', context)
    
    if response.status_code == 204:
        return redirect('manage_attribute')
    
    else:
        response_data = response.json()
        context = {
            'detail': response_data.get('detail', 'An error occurred while deleting the Attributes'),
        }
        return render(request, 'manage_attributes.html', context)

def delete_all_attributes_view(request):
    if request.method == 'POST':
        print("Welcome")
        try:
            data = json.loads(request.body)
            
            print("Data : ", data)
            
            user_ids = data.get('user_ids', [])
            
            if user_ids:
                Enquiry_Mode.objects.filter(id__in=user_ids).delete()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'No Enquiry selected for deletion'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# update_attribute_view
def update_attribute_view(request, id):
    try:
        user = Enquiry_Mode.objects.get(id=id)
    except Enquiry_Mode.DoesNotExist:
        context = {'error': 'Attributes not found'}
        return render(request, 'manage_attributes.html', context)

    if request.method == 'POST':
        
        try:
            token = Token.objects.first()  # Get the first token for simplicity
            if not token:
                raise Token.DoesNotExist
        except Token.DoesNotExist:
            context = {'error': 'Authentication token not found'}
            return render(request, 'manage_attributes.html', context)
        
        api_url = f'http://127.0.0.1:8000/api/update_enquiry_mode/{user.pk}/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }
        
        user_data = {
            'mode_of_enquiry': request.POST.get('attribute', user.mode_of_enquiry),
        }

        try:
            response = requests.patch(api_url, data=json.dumps(user_data), headers=headers)
            print("API Response Status Code:", response.status_code)
            response.raise_for_status()
            response_data = response.json()
            print("API Response Data:", response_data)
        except requests.exceptions.RequestException as err:
            context = {
                'error': f'Request error occurred: {err}',
                'response_data': response.json() if response.content else {}
            }
            return render(request, 'manage_attributes.html', context)
        
        if response.status_code in [200, 204]:  # 204 No Content is also a valid response for updates
            print("Update successful")
            return redirect('manage_attribute')
        else:
            context = {
                'error': 'Failed to update user information',
                'mode_of_enquiry': response_data.get('mode_of_enquiry', ''),
            }
            return render(request, 'update_attributes.html', context)
        
    return render(request, 'update_attributes.html', {"enquiry_mode": user})


# csv file formate for attributes
@csrf_exempt
def export_attributes_csv(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attributes_list_csv.csv"'

        writer = csv.writer(response)

        # Write the header row
        writer.writerow(['Attributes Name'])

        # Fetch selected courses based on IDs
        selected_attributes = Enquiry_Mode.objects.filter(id__in=ids)

        for user in selected_attributes:        
            # Write the data row
            writer.writerow([user.mode_of_enquiry])

        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

# Excel file format for course
@csrf_exempt
def export_attributes_excel(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Fetch selected courses based on IDs
        selected_attributes = Enquiry_Mode.objects.filter(id__in=ids)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define header row with font style and alignment
        header_row = ['Attributes Name']
        ws.append(header_row)

        for idx, user in enumerate(selected_attributes, start=2):
            ws.append([user.mode_of_enquiry])

        # Create an in-memory file-like object to save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response with Excel content type and attachment header
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
        
        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

@csrf_protect
@require_POST
def export_attributes_pdf(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        selected_attribute = Enquiry_Mode.objects.filter(id__in=ids)
        
        if not selected_attribute:
            return JsonResponse({'error': 'No users available.'}, status=404)
        
        attribute_list = []
        for user in selected_attribute:    
            attribute_list.append({
                'mode_of_enquiry': user.mode_of_enquiry, 
            })
        
        print(attribute_list)
        
        content = {'attribute_list': attribute_list}
        return renderers.render_to_pdf('attribute_data_list.html', content)
    
    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

class SearchAttributeResultsView(ListView):
    model = Enquiry_Mode
    template_name = 'search_attribute_result.html'

    def get_queryset(self):
        query = self.request.GET.get("q")
        
        object_list = Enquiry_Mode.objects.filter(
            Q(mode_of_enquiry__icontains = query)
        )
        
        return object_list

# Enquiry view

def enquiry_view(request):
    new_enquiry_no = generate_new_enquiry_no()

    if request.method == 'POST':
        try:
            inplant_no_of_days = int(request.POST.get('inplant_no_of_days', 0)) if request.POST.get('inplant_no_of_days') else None
            inplant_no_of_students = int(request.POST.get('inplant_no_of_students', 0)) if request.POST.get('inplant_no_of_students') else None
            internship_no_of_students = int(request.POST.get('internship_no_of_students', 0)) if request.POST.get('internship_no_of_students') else None
            internship_no_of_days = int(request.POST.get('internship_no_of_days', 0)) if request.POST.get('internship_no_of_days') else None
            year_of_graduation = int(request.POST.get('year_of_graduation', 0)) if request.POST.get('year_of_graduation') else None
        except ValueError:
            # Handle invalid integer or float conversion
            inplant_no_of_days = None
            inplant_no_of_students = None
            internship_no_of_students = None
            internship_no_of_days = None
            year_of_graduation = None

        # Convert date strings to the correct format (YYYY-MM-DD)
        def format_date(date_str):
            try:
                # Expecting date in dd-mm-yyyy format
                return datetime.strptime(date_str, '%d-%m-%Y').date().isoformat()
            except ValueError:
                return None  # or handle the error as needed
        
        enquiry_data = {
            'enquiry_date': format_date(request.POST.get('enquiry_date', '').strip()),
            'enquiry_no': new_enquiry_no,
            'name': request.POST.get('student_name', '').strip(),
            'contact_no': request.POST.get('contact', '').strip(),
            'email_id': request.POST.get('email', '').strip(),
            'date_of_birth': format_date(request.POST.get('dob', '').strip()),
            'fathers_name': request.POST.get('father_name', '').strip(),
            'fathers_contact_no': request.POST.get('father_contact', '').strip(),
            'fathers_occupation': request.POST.get('fathers_occupation', '').strip(),
            'address': request.POST.get('address', '').strip(),
            'status': request.POST.get('status', '').strip(),
            'course_name': request.POST.get('course_name', '').strip(),
            'inplant_technology': request.POST.get('inplant_technology', '').strip(),
            'inplant_no_of_days': inplant_no_of_days,
            'inplant_no_of_students': inplant_no_of_students,
            'internship_technology': request.POST.get('internship_technology', '').strip(),
            'internship_no_of_days': internship_no_of_days,
            'internship_no_of_students': internship_no_of_students,
            'next_follow_up_date': format_date(request.POST.get('next_follow_up_date', '').strip()),
            'degree': request.POST.get('degree', '').strip(),
            'college': request.POST.get('college', '').strip(),
            'grade_percentage': request.POST.get('grade_percentage', '').strip(),
            'year_of_graduation': year_of_graduation,
            'mode_of_enquiry': request.POST.get('mode_of_enquiry', '').strip(),
            'reference_name': request.POST.get('reference_name', '').strip(),
            'reference_contact_no': request.POST.get('reference_contact', '').strip(),
            'other_enquiry_details': request.POST.get('other_enquiry_details', '').strip(),
            'lead_type': request.POST.get('lead_type', '').strip(),
        }

        # Get the token
        try:
            token = Token.objects.get(user=request.user)
        except Token.DoesNotExist:
            context = {
                'error': 'Authentication token not found',
                **enquiry_data,
            }
            return render(request, 'new_enquiry.html', context)
        
        api_url = 'http://127.0.0.1:8000/api/enquiry/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json'
        }

        try:
            response = requests.post(api_url, json=enquiry_data, headers=headers)
            response_data = response.json()
            print(response_data)
            
        except requests.exceptions.RequestException:
            context = {
                'error': 'An error occurred while creating the enquiry.',
                **enquiry_data,
            }
            return render(request, 'new_enquiry.html', context)        
        
        if response.status_code == 201:
            messages.success(request, 'Created Successfully')
            return redirect('enquiry')
        else:
            # Handle API error messages
            error_message = response_data.get('error', 'An error occurred during enquiry creation.')
            errors = response_data
            
            print(errors)
            
            courses = Course.objects.all()
            mode_of_enquiry_choices = Enquiry_Mode.objects.all()
            context = {
                'error': error_message,
                'errors': errors,
                **enquiry_data,
                'courses': courses,
                'mode_of_enquiry_choices': mode_of_enquiry_choices,
            }
            return render(request, 'new_enquiry.html', context)
    
    courses = Course.objects.all()
    mode_of_enquiry_choices = Enquiry_Mode.objects.all()

    context = {
        'courses': courses,
        'mode_of_enquiry_choices': mode_of_enquiry_choices,
        'enquiry_no': new_enquiry_no
    }
    
    return render(request, 'new_enquiry.html', context)


def manage_enquiry_view(request):
    # Fetch the token
    
    enquiry_data = Enquiry_Mode.objects.all().values()
    
    course = Course.objects.all().values()
    
    try:
        token = Token.objects.get(user=request.user)  # Assuming you only have one token and it's safe to get the first one
    except Token.DoesNotExist:
        context = {
            'error': 'Authentication token not found'
        }
        return render(request, 'manage_user.html', context)
    
    api_url = 'http://127.0.0.1:8000/api/enquiry/'
    
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        response_data = response.json()
        
    except requests.exceptions.RequestException as err:
        # Catch any request-related exceptions
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_user.html', context)

    # Get the per_page value from the request, default to 10 if not provided
    per_page = request.GET.get('per_page', '10')

    # Apply pagination
    paginator = Paginator(response_data, per_page)  # Use response_data for pagination
    
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
        'course_name': course,
        'enquiry': enquiry_data,
    }
    return render(request, 'manage_enquiry.html', context)

def update_enquiry_view(request, id):
    try:
        enquiry = Enquiry.objects.get(id=id)
    except Enquiry.DoesNotExist:
        context = {'error': 'Enquiry not found'}
        return render(request, 'manage_enquiry.html', context)

    if request.method == 'POST':
        print("Welcome")
        
        try:
            token = Token.objects.get(user=request.user)  # Get the first token for simplicity
            print(token)
            if not token:
                raise Token.DoesNotExist
        except Token.DoesNotExist:
            context = {'error': 'Authentication token not found'}
            return render(request, 'manage_enquiry.html', context)
        
        api_url = f'http://127.0.0.1:8000/api/update_enquiry/{enquiry.pk}/'
        
        print("Api URL : ", api_url)
        
        headers = {
            'Authorization': f'Token {token.key}',
        }
        
        # Prepare the data
        enquiry_data = {
            'enquiry_date': request.POST.get('enquiry_date', enquiry.enquiry_date),
            'enquiry_no': request.POST.get('enquiry_no', enquiry.enquiry_no),
            'name': request.POST.get('student_name', enquiry.name),
            'contact_no': request.POST.get('contact', enquiry.contact_no),
            'email_id': request.POST.get('email', enquiry.email_id),
            'date_of_birth': request.POST.get('dob', enquiry.date_of_birth),
            'fathers_name': request.POST.get('father_name', enquiry.fathers_name),
            'fathers_contact_no': request.POST.get('father_contact', enquiry.fathers_contact_no),
            'fathers_occupation': request.POST.get('fathers_occupation', enquiry.fathers_occupation),
            'address': request.POST.get('address', enquiry.address),
            # 'status': request.POST.get('status', enquiry.status),
            'course_name': request.POST.get('course_name', enquiry.course_name),
            'inplant_technology': request.POST.get('technology', enquiry.inplant_technology),
            'inplant_no_of_days': request.POST.get('inplant_no_of_days', enquiry.inplant_no_of_days),
            'inplant_no_of_students': request.POST.get('inplant_no_of_students', enquiry.inplant_no_of_students),
            'internship_technology': request.POST.get('internship_technology', enquiry.internship_technology),
            'internship_no_of_days': request.POST.get('internship_no_of_days', enquiry.internship_no_of_days),
            'next_follow_up_date': request.POST.get('next_follow_up_date', enquiry.next_follow_up_date),
            'degree': request.POST.get('degree', enquiry.degree),
            'college': request.POST.get('college', enquiry.college),
            'grade_percentage': request.POST.get('grade_percentage', enquiry.grade_percentage),
            'year_of_graduation': request.POST.get('year_of_graduation', enquiry.year_of_graduation),
            'mode_of_enquiry': request.POST.get('mode_of_enquiry', enquiry.mode_of_enquiry),
            'reference_name': request.POST.get('reference_name', enquiry.reference_name),
            'reference_contact_no': request.POST.get('reference_contact', enquiry.reference_contact_no),
            'other_enquiry_details': request.POST.get('other_enquiry_details', enquiry.other_enquiry_details),
            'notes': request.POST.get('notes', enquiry.notes),
            'status': request.POST.get('notes', enquiry.notes),
            'lead_type': request.POST.get('lead_type', enquiry.lead_type),
        }
        
        files = {}
        if 'files' in request.FILES:
            files = {'files': request.FILES['files']}
        else:
            files = {'files': enquiry.files} # Or handle the case when no new file is uploaded
            
        # Log the data and files to be sent
        print("Enquiry Data:", enquiry_data)
        print("Files Data:", files)
        
        try:
            response = requests.patch(api_url, data=enquiry_data, files=files, headers=headers)
            print("API Response Status Code:", response.status_code)
            response.raise_for_status()
            response_data = response.json()
            print("API Response Data:", response_data)
        except requests.exceptions.RequestException as err:
            print(f'Request error occurred: {err}')
            context = {
                'error': f'Request error occurred: {err}',
                'response_data': response.json() if response.content else {}
            }
            return render(request, 'manage_enquiry.html', context)
        
        if response.status_code in [200, 204]:  # 204 No Content is also a valid response for updates
            # notes_content = request.POST.get('notes', '')
            # uploaded_file = request.FILES.get('files', None)
            notes_content = request.POST.get('notes', '')
            uploaded_file = request.FILES.get('files', None)
            
            print("Notes Content : ", notes_content)
            
            print("Notes File : ", uploaded_file)
            
            if notes_content or uploaded_file:
                notes_data = Notes.objects.create(
                    notes=notes_content,
                    files=uploaded_file,
                    user_id = enquiry.pk,
                    created_by=request.user.id,  # Assuming the request user is authenticated
                    modified_by=request.user.id
                )
                
                print(notes_data)
                
                notes_data.save()
            print("Update successful")
            return redirect('manage_enquiry')
        else:
            context = {
                'error': response_data.get('error', 'An error occurred during enquiry creation.'),
                'enquiry_data': enquiry_data,
                'files': files.get('files', ''),
            }
            return render(request, 'update_enquiry.html', context)
    
    courses = Course.objects.all()
    notes = Notes.objects.all().values().order_by("-created_at")
    mode_of_enquiry = Enquiry_Mode.objects.all()
    
    context = {
        'courses': courses,
        'mode_of_enquiry': mode_of_enquiry,
        "enquiry": enquiry,
        "notes": notes,
        "enquiry_id": id,
    }    
        
    return render(request, 'update_enquiry.html', context)

def delete_enquiry_view(request, id):
    user_id = Enquiry.objects.get(id=id)
    
    print(user_id.pk)
    
    if not user_id:
        context = {'error': 'Enquiry ID not provided'}
        return render(request, 'manage_enquiry.html', context)
    
    try:
        token = Token.objects.get(user=request.user)  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_enquiry.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/delete_enquiry/{user_id.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()
        
        if response.status_code == 200:
            messages.success(request, 'Successfully Deleted')
            return redirect('manage_enquiry')

    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_enquiry.html', context)
    
def delete_all_enquiry_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            print("Data : ", data)
            
            user_ids = data.get('user_ids', [])
            
            print("User ID : ", user_ids)
            
            if user_ids:
                Enquiry.objects.filter(id__in=user_ids).delete()
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'No users selected for deletion'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# csv file formate for attributes
@csrf_exempt
def export_enquiry_csv(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="attributes_list_csv.csv"'

        writer = csv.writer(response)

        # Write the header row with capitalized first letters
        writer.writerow([
            'Enquiry Date', 'Enquiry No', 'Name', 'Contact No', 'Course Name', 'Next Follow Up Date', 
            'Mode of Enquiry', 'Lead Type', 'Status'
        ])

        # Fetch selected enquiries based on IDs
        selected_enquiries = Enquiry.objects.filter(id__in=ids)

        for enquiry in selected_enquiries:
            writer.writerow([
                enquiry.enquiry_date,
                enquiry.enquiry_no,
                enquiry.name,
                int(enquiry.contact_no),
                enquiry.course_name.course_name if enquiry.course_name else '',  # Use the course name
                enquiry.next_follow_up_date,
                enquiry.mode_of_enquiry.mode_of_enquiry if enquiry.mode_of_enquiry else '',  # Use the mode of enquiry name
                enquiry.lead_type,
                enquiry.status
            ])

        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

# Excel file format for course
@csrf_exempt
def export_enquiry_excel(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Fetch selected courses based on IDs
        selected_attributes = Enquiry.objects.filter(id__in=ids)
        
        if not selected_attributes:
            return JsonResponse({'error': 'No Enquires available.'}, status=404)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define header row with font style and alignment
        # Define header row with capitalized first letters
        headers = [
            'Enquiry Date', 'Enquiry No', 'Name', 'Contact No',
            'Course Name', 'Next Follow Up Date',
            'Mode of Enquiry', 'Lead Type', 'Status'
        ]
        
        # Append the header row to the sheet
        ws.append(headers)

        for enquiry in selected_attributes:
            ws.append([
                enquiry.enquiry_date.strftime('%Y-%m-%d') if enquiry.enquiry_date else '',
                enquiry.enquiry_no,
                enquiry.name,
                int(enquiry.contact_no),
                enquiry.course_name.course_name if enquiry.course_name else '',
                enquiry.next_follow_up_date.strftime('%Y-%m-%d') if enquiry.next_follow_up_date else '',
                enquiry.mode_of_enquiry.mode_of_enquiry if enquiry.mode_of_enquiry else '',
                enquiry.lead_type,
                enquiry.status
            ])

        # Create an in-memory file-like object to save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response with Excel content type and attachment header
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
        
        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

@csrf_protect
@require_POST
def export_enquiry_pdf(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        selected_attribute = Enquiry.objects.filter(id__in=ids)
        
        if not selected_attribute:
            return JsonResponse({'error': 'No users available.'}, status=404)
        
        attribute_list = []
        for enquiry in selected_attribute:    
            attribute_list.append({
                'enquiry_date': enquiry.enquiry_date.strftime('%Y-%m-%d') if enquiry.enquiry_date else '',
                'enquiry_no': enquiry.enquiry_no,
                'name': enquiry.name,
                'contact_no': enquiry.contact_no,
                'email_id': enquiry.email_id,
                'date_of_birth': enquiry.date_of_birth.strftime('%Y-%m-%d') if enquiry.date_of_birth else '',
                'fathers_name': enquiry.fathers_name,
                'fathers_contact_no': enquiry.fathers_contact_no,
                'fathers_occupation': enquiry.fathers_occupation,
                'address': enquiry.address,
                'status': enquiry.status,
                'course_name': enquiry.course_name.course_name if enquiry.course_name else '',
                'inplant_technology': enquiry.inplant_technology,
                'inplant_no_of_days': enquiry.inplant_no_of_days if enquiry.inplant_no_of_days is not None else '',
                'inplant_no_of_students': enquiry.inplant_no_of_students if enquiry.inplant_no_of_students is not None else '',
                'internship_technology': enquiry.internship_technology,
                'internship_no_of_days': enquiry.internship_no_of_days if enquiry.internship_no_of_days is not None else '',
                'internship_no_of_students': enquiry.internship_no_of_students if enquiry.internship_no_of_students is not None else '',
                'next_follow_up_date': enquiry.next_follow_up_date.strftime('%Y-%m-%d') if enquiry.next_follow_up_date else '',
                'degree': enquiry.degree,
                'college': enquiry.college,
                'grade_percentage': enquiry.grade_percentage if enquiry.grade_percentage is not None else '',
                'year_of_graduation': enquiry.year_of_graduation if enquiry.year_of_graduation is not None else '',
                'mode_of_enquiry': enquiry.mode_of_enquiry.mode_of_enquiry if enquiry.mode_of_enquiry else '',
                'reference_name': enquiry.reference_name,
                'reference_contact_no': enquiry.reference_contact_no,
                'other_enquiry_details': enquiry.other_enquiry_details,
                'lead_type': enquiry.lead_type
            })        
        content = {'enquiry_list': attribute_list}
        return renderers.render_to_pdf('enquiry_data_list.html', content)
    
    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

class SearchEnquiryResultsView(ListView):
    model = Enquiry
    template_name = 'search_enquiry_result.html'

    def get_queryset(self):
        query = self.request.GET.get("q", "")
        start_date = self.request.GET.get("start_date", "")
        end_date = self.request.GET.get("end_date", "")

        object_list = Enquiry.objects.all()

        # Apply text search if query is provided
        if query:
            object_list = object_list.filter(
                Q(enquiry_no__icontains=query) |
                Q(name__icontains=query) |
                Q(contact_no__icontains=query) |
                Q(course_name__course_name__icontains=query) |
                Q(mode_of_enquiry__mode_of_enquiry__icontains=query) |
                Q(status__icontains=query) |
                Q(lead_type__icontains=query)
            )

        # Apply date range filter if dates are provided
        if start_date and end_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                object_list = object_list.filter(enquiry_date__range=(start_date, end_date))
            except ValueError:
                pass  # Handle invalid date format if necessary

        return object_list
    
    
def delete_notes_view(request, id):
    user_id = Notes.objects.get(id=id)
    
    print(user_id)
    
    if not user_id:
        context = {'error': 'Attribute ID not provided'}
        return render(request, 'update_enquiry.html', context)
    
    try:
        token = Token.objects.get(user=request.user)  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'update_enquiry.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/delete_notes/{user_id.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()
        
        if response.status_code == 200:
            return redirect('manage_enquiry')
    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'update_enquiry.html', context)
    

def update_notes_view(request, id):
    try:
        user = Notes.objects.get(id=id)
    except Notes.DoesNotExist:
        context = {'error': 'Notes not found'}
        return render(request, 'update_enquiry.html', context)

    if request.method == 'POST':
        try:
            token = Token.objects.get(user=request.user)
            if not token:
                raise Token.DoesNotExist
        except Token.DoesNotExist:
            context = {'error': 'Authentication token not found'}
            return render(request, 'update_enquiry.html', context)
        
        api_url = f'http://127.0.0.1:8000/api/update_notes/{user.pk}/'
        headers = {
            'Authorization': f'Token {token.key}',
        }
        
        files = {}
        if 'files' in request.FILES:
            files['files'] = request.FILES['files']
        
        user_data = {
            'notes': request.POST.get('notes', user.notes),
        }

        try:
            response = requests.patch(api_url, data=user_data, files=files, headers=headers)
            response.raise_for_status()
            response_data = response.json()
        except requests.exceptions.RequestException as err:
            context = {
                'error': f'Request error occurred: {err}',
                'response_data': response.json() if response.content else {}
            }
            return render(request, 'update_enquiry.html', context)
        
        if response.status_code == 200:  # 204 No Content is also a valid response for updates
            print("Update successful")
            return redirect('manage_enquiry')
        else:
            context = {
                'error': 'Failed to update user information',
                'notes': response_data.get('notes', ''),
            }
            return render(request, 'update_enquiry.html', context)
        
    return render(request, 'update_notes.html')


@require_GET
def get_enquiry_details(request):
    enquiry_no = request.GET.get('enquiry_no')
    
    print("Enquiry No : ", enquiry_no)
    
    if not enquiry_no:
        return JsonResponse({'error': 'No enquiry number provided'}, status=400)
    
    try:
        enquiry = Enquiry.objects.get(enquiry_no=enquiry_no)
        
        data = {
            'name': enquiry.name,
            'contact_no': enquiry.contact_no,
            'date_of_birth': enquiry.date_of_birth.strftime('%Y-%m-%d') if enquiry.date_of_birth else '',
            'email_id': enquiry.email_id,
            'fathers_name': enquiry.fathers_name,
            'fathers_contact_no': enquiry.fathers_contact_no,
            'degree': enquiry.degree,
            'grade_percentage': enquiry.grade_percentage,
            'year_of_graduation': enquiry.year_of_graduation,
            'college': enquiry.college,
            'course_name': enquiry.course_name.course_name,  # Assuming course_name is a ForeignKey
            'inplant_technology': enquiry.inplant_technology,
            'inplant_no_of_days': enquiry.inplant_no_of_days,
            'inplant_no_of_students': enquiry.inplant_no_of_students,
            'internship_technology': enquiry.internship_technology,
            'internship_no_of_days': enquiry.internship_no_of_days,
            'internship_no_of_students': enquiry.internship_no_of_students,
        }
        
        print("Data : ", data)
        
        return JsonResponse(data, status=200)
    
    except Enquiry.DoesNotExist:
        return JsonResponse({'error': 'Enquiry not found'}, status=404)

def new_enrollment_view(request):
    new_registration_no = generate_new_registration_no()
    
    print("Registration Number : ", new_registration_no)
    
    try:
        inplant_no_of_days = int(request.POST.get('inplant_no_of_days', 0)) if request.POST.get('inplant_no_of_days') else None
        inplant_no_of_students = int(request.POST.get('inplant_no_of_students', 0)) if request.POST.get('inplant_no_of_students') else None
        internship_no_of_students = int(request.POST.get('internship_no_of_students', 0)) if request.POST.get('internship_no_of_students') else None
        internship_no_of_days = int(request.POST.get('internship_no_of_days', 0)) if request.POST.get('internship_no_of_days') else None
    except ValueError as e:
        print("ValueError during conversion:", e)
        inplant_no_of_days = None
        inplant_no_of_students = None
        internship_no_of_students = None
        internship_no_of_days = None
    
    if request.method == 'POST':
        enquiry_no = request.POST.get('enquiry_no')
        
        # Fetch the related Enquiry object
        try:
            enquiry = Enquiry.objects.get(enquiry_no=enquiry_no)
        except Enquiry.DoesNotExist:
            context = {
                'error': 'Enquiry with the provided Enquiry Number does not exist.',
            }
            return render(request, 'new_enrollment.html', context)
        
        # Check if the enquiry already exists
        if Enrollment.objects.filter(enquiry_no=enquiry_no).exists():
            context = {
                'error': 'This Enquiry Number Already Exist.',
            }
            return render(request, 'new_enrollment.html', context)

        # Ensure that grade_percentage is not None
        grade_percentage = enquiry.grade_percentage if enquiry.grade_percentage is not None else request.POST.get('grade_percentage')

        print("Grade Percentage : ", grade_percentage)
        
        # Format the registration_date
        try:
            registration_date_str = request.POST.get('registration_date')
            if registration_date_str:
                registration_date = datetime.strptime(registration_date_str, '%d-%m-%Y').strftime('%Y-%m-%d')
            else:
                registration_date = None
        except ValueError:
            registration_date = None
            print("ValueError: Invalid date format for registration_date")

        # Auto-populate fields based on the related Enquiry object
        enrollment_data = {
            'enquiry_no': enquiry.enquiry_no,
            'registration_no': new_registration_no,
            'registration_date': registration_date,
            'name': enquiry.name,
            'phonenumber': enquiry.contact_no,
            'date_of_birth': enquiry.date_of_birth.strftime('%Y-%m-%d') if enquiry.date_of_birth else '',
            'gender': request.POST.get('gender'),
            'email_id': enquiry.email_id,
            'father_name': enquiry.fathers_name,
            'fathers_email_id': request.POST.get('fathers_email_id'),
            'fathers_contact_no': enquiry.fathers_contact_no,
            'degree': enquiry.degree,
            'institution': request.POST.get('institution'),
            'subject': request.POST.get('subject'),
            'grade_percentage': grade_percentage,
            'year_of_passed_out': enquiry.year_of_graduation,
            'place': request.POST.get('place'),
            'designation': request.POST.get('designation'),
            'company_name': request.POST.get('company_name'),
            'work_experience': request.POST.get('work_experience'),
            'nature_of_work': request.POST.get('nature_of_work'),
            # 'from_date': request.POST.get('from_date'),
            # 'to_date': request.POST.get('to_date'),
            'course_name': enquiry.course_name.id,
            'inplant_technology': request.POST.get('inplant_technology', '').strip(),
            'inplant_no_of_days': inplant_no_of_days,
            'inplant_no_of_students': inplant_no_of_students,
            'internship_technology': request.POST.get('internship_technology', '').strip(),
            'internship_no_of_days': internship_no_of_days,
            'internship_no_of_students': internship_no_of_students,
            'duration': request.POST.get('duration'),
            'payment_type': request.POST.get('payment_type'),
            'total_fees_amount': request.POST.get('total_fees_amount'),
            'installment_amount': request.POST.get('installment_amount'),
        }

        print("Enrollment Data Being Sent:", json.dumps(enrollment_data, indent=4))  # Log the data

        try:
            token = Token.objects.get(user=request.user)
        except Token.DoesNotExist:
            context = {
                'error': 'Authentication Token not Found',
                **enrollment_data,
            }
            return render(request, 'new_enrollment.html', context)

        api_url = 'http://127.0.0.1:8000/api/enrollment/'
        
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json',
        }

        try:
            response = requests.post(api_url, json=enrollment_data, headers=headers)
            response_data = response.json()

            print("API Response:", response_data)  # Log the response

            if response.status_code in [200, 201]:
                messages.success(request, 'Created Successfully')
                return redirect('enrollment')  # Redirect to a success page or another view
            else:
                error_message = response_data.get('error', 'An Error Occurred During Creation.')
                errors = response_data
                courses = Course.objects.all()
                context = {
                    'error': error_message,
                    'errors': errors,
                    'courses': courses,
                    **enrollment_data,
                }
                return render(request, 'new_enrollment.html', context)

        except requests.exceptions.RequestException as e:
            print("RequestException occurred:", e)
            context = {
                'error': 'An Error Occurred While Creating an Enrollment',
                **enrollment_data,
            }
            return render(request, 'new_enrollment.html', context)

    courses = Course.objects.all()
    
    context = {
        'courses': courses,
        'registration_no': new_registration_no,
    }
    
    return render(request, 'new_enrollment.html', context)


def manage_enrollment_view(request):
    
    course = Course.objects.all().values()
    
    try:
        token = Token.objects.get(user=request.user)  # Assuming you only have one token and it's safe to get the first one
    except Token.DoesNotExist:
        context = {
            'error': 'Authentication token not found'
        }
        return render(request, 'manage_enrollment.html', context)
    
    api_url = 'http://127.0.0.1:8000/api/enrollment/'
    
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()  # Raise an HTTPError for bad responses
        response_data = response.json()
        
    except requests.exceptions.RequestException as err:
        # Catch any request-related exceptions
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_enrollment.html', context)

    # Get the per_page value from the request, default to 10 if not provided
    per_page = request.GET.get('per_page', '10')

    # Apply pagination
    paginator = Paginator(response_data, per_page)  # Use response_data for pagination
    
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    context = {
        'page_obj': page_obj,
        'per_page': per_page,
        'course_name': course,
    }
    return render(request, 'manage_enrollment.html', context)

def update_enrollment_view(request, id):
    try:
        enrollment = Enrollment.objects.get(id=id)
    except Enrollment.DoesNotExist:
        context = {'error': 'Enrollment not found'}
        return render(request, 'manage_enrollment.html', context)

    if request.method == 'POST':
        
        try:
            token = Token.objects.get(user=request.user)  # Get the first token for simplicity
            if not token:
                raise Token.DoesNotExist
        except Token.DoesNotExist:
            context = {'error': 'Authentication token not found'}
            return render(request, 'manage_enrollment.html', context)
        
        api_url = f'http://127.0.0.1:8000/api/update_enrollment/{enrollment.pk}/'
        
        headers = {
            'Authorization': f'Token {token.key}',
        }
        
        # Get the registration date from the POST request
        input_registration_date = request.POST.get('registration_date', enrollment.registration_date.strftime("%Y-%m-%d"))
        print(f"Raw registration_date input: {input_registration_date}")

        # Try to parse the registration date
        try:
            parsed_registration_date = datetime.strptime(input_registration_date, "%d-%m-%Y")  # Adjust if input format changes
            registration_date = parsed_registration_date.strftime("%Y-%m-%d")
        except ValueError:
            registration_date = enrollment.registration_date.strftime("%Y-%m-%d")

        print(f"Formatted registration_date: {registration_date}")

        # Get the date of birth from the POST request
        input_date_of_birth = request.POST.get('date_of_birth', enrollment.date_of_birth.strftime("%Y-%m-%d"))
        print(f"Raw date_of_birth input: {input_date_of_birth}")

        # Try to parse the date of birth
        try:
            parsed_date_of_birth = datetime.strptime(input_date_of_birth, "%d-%m-%Y")  # Adjust if input format changes
            date_of_birth = parsed_date_of_birth.strftime("%Y-%m-%d")
        except ValueError:
            date_of_birth = enrollment.date_of_birth.strftime("%Y-%m-%d")

        print(f"Formatted date_of_birth: {date_of_birth}")

        # Auto-populate fields based on the related Enquiry object
        enrollment_data = {
            'enquiry_no': request.POST.get('enquiry_no', enrollment.enquiry_no),
            'registration_no': request.POST.get('registration_no', enrollment.registration_no),
            'registration_date': registration_date,
            'name': request.POST.get('name', enrollment.name),
            'phonenumber': request.POST.get('phonenumber', enrollment.phonenumber),
            'date_of_birth': date_of_birth,
            'gender': request.POST.get('gender', enrollment.gender),
            'email_id': request.POST.get('email_id', enrollment.email_id),
            'father_name': request.POST.get('father_name', enrollment.father_name),
            'fathers_email_id': request.POST.get('father_email_id', enrollment.fathers_email_id),
            'fathers_contact_no': request.POST.get('fathers_contact_no', enrollment.fathers_contact_no),
            'degree': request.POST.get('degree', enrollment.degree),
            'institution': request.POST.get('institution', enrollment.institution),
            'subject': request.POST.get('subject', enrollment.subject),
            'grade_percentage' : request.POST.get('grade_percentage', enrollment.grade_percentage),
            'place' : request.POST.get('place', enrollment.place),
            'year_of_passed_out': request.POST.get('year_of_passed_out', enrollment.year_of_passed_out),
            'designation': request.POST.get('designation', enrollment.designation),
            'company_name': request.POST.get('company_name', enrollment.company_name),
            'work_experience': request.POST.get('work_experience', enrollment.work_experience),
            'nature_of_work': request.POST.get('nature_of_work', enrollment.work_experience),
            'course_name': request.POST.get('course_name', enrollment.course_name),
            'inplant_technology': request.POST.get('inplant_technology', enrollment.inplant_technology),
            'inplant_no_of_days': request.POST.get('inplant_no_of_days', enrollment.inplant_no_of_days),
            'inplant_no_of_students': request.POST.get('inplant_no_of_students', enrollment.inplant_no_of_students),
            'internship_technology': request.POST.get('internship_technology', enrollment.internship_technology),
            'internship_no_of_days': request.POST.get('internship_no_of_days', enrollment.internship_no_of_days),
            'internship_no_of_students': request.POST.get('internship_no_of_students', enrollment.internship_no_of_students),
            'duration': request.POST.get('duration', enrollment.duration),
            'payment_type': request.POST.get('payment_type', enrollment.payment_type),
            'total_fees_amount': request.POST.get('total_fees_amount', enrollment.total_fees_amount),
            'installment_amount': request.POST.get('installment_amount', enrollment.installment_amount),
        }
        
        print("Enrollment Data : ", enrollment_data)
        
        try:
            response = requests.patch(api_url, data=enrollment_data, headers=headers)
            print("API Response Status Code:", response.status_code)
            response.raise_for_status()
            response_data = response.json()
            print("API Response Data:", response_data)
        except requests.exceptions.RequestException as err:
            print(f'Request error occurred: {err}')
            context = {
                'error': f'Request error occurred: {err}',
                'response_data': response.json() if response.content else {}
            }
            return render(request, 'manage_enrollment.html', context)
        
        if response.status_code in [200, 204]:  # 204 No Content is also a valid response for updates
            messages.success(request, 'Successfully Updated')
            return redirect('manage_enrollment')
        else:
            context = {
                'error': response_data.get('error', 'An error occurred during enquiry creation.'),
                'enrollment_data': enrollment_data,
            }
            return render(request, 'update_enrollment.html', context)
    
    courses = Course.objects.all()
    reg_date = enrollment.registration_date.strftime("%d-%m-%Y")
    context = {
        'courses': courses,
        "enrollment": enrollment,
        "reg_date": reg_date,
        "enquiry_id": id,
    }
    
    print("Registration Date : ", enrollment.registration_date.strftime("%d-%m-%Y"))
        
    return render(request, 'update_enrollment.html', context)

def delete_enrollment_view(request, id):
    user_id = Enrollment.objects.get(id=id)
    
    if not user_id:
        context = {'error': 'Enrollment ID not provided'}
        return render(request, 'manage_enrollment.html', context)
    
    try:
        token = Token.objects.get(user=request.user)  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_enquiry.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/delete_enrollment/{user_id.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()
        
        if response.status_code == 200:
            messages.success(request, 'Successfully Deleted')
            return redirect('manage_enrollment')

    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_enrollment.html', context)
    
def delete_all_enrollment_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            print("Data : ", data)
            
            user_ids = data.get('user_ids', [])
            
            print("User ID : ", user_ids)
            
            if user_ids:
                Enrollment.objects.filter(id__in=user_ids).delete()
                messages.success(request, 'Successfully Deleted')
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'No users selected for deletion'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

# csv file formate for attributes
@csrf_exempt
def export_enrollment_csv(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Create the HttpResponse object with the appropriate CSV header
        response = HttpResponse(content_type='text/csv')
        response['Content-Disposition'] = 'attachment; filename="enrollment_list_csv.csv"'

        writer = csv.writer(response)

        # Write the header row with capitalized first letters
        
        writer.writerow([
            'Registratio No', 'Registratio Date', 'Name', 'Contact No', 'Course Name', 'Course Duration', 
            'Total Fees',
        ])

        # Fetch selected enquiries based on IDs
        selected_enrollments = Enrollment.objects.filter(id__in=ids)

        for enrollment in selected_enrollments:
            writer.writerow([
                enrollment.registration_no,
                enrollment.registration_date,
                enrollment.name,
                int(enrollment.phonenumber),
                enrollment.course_name.course_name if enrollment.course_name else '',  # Use the course name
                enrollment.duration,
                enrollment.total_fees_amount,
            ])

        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

# Excel file format for course
@csrf_exempt
def export_enrollment_excel(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

        # Fetch selected courses based on IDs
        selected_enrollments = Enrollment.objects.filter(id__in=ids)
        
        if not selected_enrollments:
            return JsonResponse({'error': 'No Enrollment available.'}, status=404)

        # Create an Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active

        # Define header row with font style and alignment
        # Define header row with capitalized first letters
        headers = [
            'Registratio No', 'Registratio Date', 'Name', 'Contact No', 'Course Name', 'Course Duration', 
            'Total Fees',
        ]
        
        # Append the header row to the sheet
        ws.append(headers)

        for enrollment in selected_enrollments:
            ws.append([
                enrollment.registration_no,
                enrollment.registration_date,
                enrollment.name,
                int(enrollment.phonenumber),
                enrollment.course_name.course_name if enrollment.course_name else '',  # Use the course name
                int(enrollment.duration),
                enrollment.total_fees_amount,
            ])

        # Create an in-memory file-like object to save the workbook
        output = BytesIO()
        wb.save(output)
        output.seek(0)

        # Create the HTTP response with Excel content type and attachment header
        response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = 'attachment; filename="generated_excel.xlsx"'
        
        return response

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

@csrf_protect
@require_POST
def export_enrollment_pdf(request):
    if request.method == 'POST':
        ids = request.POST.get('ids', '').split(',')
        selected_enrollments = Enrollment.objects.filter(id__in=ids)
        
        if not selected_enrollments:
            return JsonResponse({'error': 'No users available.'}, status=404)
        
        attribute_list = []
        for enrollment in selected_enrollments:    
            attribute_list.append({
                'registration_no': enrollment.registration_no,
                'registration_date': enrollment.registration_date,
                'name':enrollment.name,
                'phonenumber': enrollment.phonenumber,
                'course_name': enrollment.course_name.course_name,
                'duration': enrollment.duration,
                'total_fees_amount': enrollment.total_fees_amount,
            })        
        content = {'enrollment_list': attribute_list}
        return renderers.render_to_pdf('enrollment_data_list.html', content)
    
    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

class SearchEnrollmentResultsView(ListView):
    model = Enrollment
    template_name = 'search_enrollment_result.html'

    def get_queryset(self):
        start_date = self.request.GET.get("start_date", "")
        end_date = self.request.GET.get("end_date", "")
        query = self.request.GET.get("q", "")

        # Start with all enrollments
        object_list = Enrollment.objects.all()

        # Apply text search if query is provided
        if query:
            object_list = object_list.filter(
                Q(registration_no__icontains=query) |
                Q(name__icontains=query) |
                Q(phonenumber__icontains=query) |
                Q(course_name__course_name__icontains=query) |
                Q(duration__icontains=query) |
                Q(total_fees_amount__icontains=query)
            )
            
        # Apply date range filter if dates are provided
        if start_date and end_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                object_list = object_list.filter(registration_date__range=(start_date, end_date))
                
            except ValueError:
                messages.add_message(self.request, messages.ERROR, "Invalid date format. Please use YYYY-MM-DD.")
        
        # Optimize the query by selecting related course_name objects
        object_list = object_list.select_related('course_name')

        return object_list


@require_GET
def get_enrollment_details(request):
    registration_no = request.GET.get('registration_no')
    
    print("Registration No : ", registration_no)
    
    if not registration_no:
        return JsonResponse({'error': 'No enquiry number provided'}, status=400)
    
    try:
        enrollment = Enrollment.objects.get(registration_no=registration_no)
        
        data = {
            'student_name': enrollment.name,
            'course_name': enrollment.course_name.course_name,
            'duration': enrollment.duration,
            'inplant_no_of_days': enrollment.inplant_no_of_days,
            'internship_no_of_days': enrollment.internship_no_of_days,
            'joining_date': enrollment.registration_date.strftime('%Y-%m-%d') if enrollment.registration_date else '',
            'total_fees': enrollment.total_fees_amount,
            'installment_amount': enrollment.installment_amount,
            'payment_type': enrollment.payment_type,
        }
        
        print("Data : ", data)
        
        return JsonResponse(data, status=200)
    
    except Enquiry.DoesNotExist:
        return JsonResponse({'error': 'Enquiry not found'}, status=404)

# new payment view 

def validate_date(date_str):
    """Validate and convert date string to datetime.date object."""
    if date_str:
        try:
            return datetime.strptime(date_str, '%Y-%m-%d').date()
        except ValueError:
            raise ValidationError("Date has wrong format. Use YYYY-MM-DD.")
    return None

def delete_payment_view(request, id):
    user_id = PaymentInfo.objects.get(id=id)
    
    if not user_id:
        context = {'error': 'Payment ID not provided'}
        return render(request, 'manage_payment_info.html', context)
    
    try:
        token = Token.objects.get(user=request.user)  # Get the first token for simplicity
        if not token:
            raise Token.DoesNotExist
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_payment_info.html', context)
    
    api_url = f'http://127.0.0.1:8000/api/delete_payment_info/{user_id.pk}/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }
    
    try:
        response = requests.delete(api_url, headers=headers)
        response.raise_for_status()
        
        if response.status_code == 200:
            messages.success(request, 'Successfully Deleted')
            return redirect('new_manage_payments')

    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': response.json() if response else {}
        }
        return render(request, 'manage_payment_info.html', context)

def delete_all_payment_view(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            
            print("Data : ", data)
            
            user_ids = data.get('user_ids', [])
            
            print("User ID : ", user_ids)
            
            if user_ids:
                PaymentInfo.objects.filter(id__in=user_ids).delete()
                messages.success(request, 'Successfully Deleted')
                return JsonResponse({'success': True})
            else:
                return JsonResponse({'success': False, 'error': 'No users selected for deletion'})
        except json.JSONDecodeError:
            return JsonResponse({'success': False, 'error': 'Invalid JSON'})
    return JsonResponse({'success': False, 'error': 'Invalid request method'})

@csrf_exempt
def export_payment_csv(request):
    if request.method == 'POST':
        try:
            ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

            # Fetch selected payments based on IDs
            selected_payments = PaymentInfo.objects.filter(id__in=ids)

            if not selected_payments:
                logging.error("No Payment available for the provided IDs.")
                return JsonResponse({'error': 'No Payment available.'}, status=404)

            # Create the HttpResponse object with the appropriate CSV header
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="payment_list_csv.csv"'

            writer = csv.writer(response)

            # Write the header row
            writer.writerow([
                'Registration No', 'Joining Date', 'Student Name', 'Course Name', 'Course Duration',
                'Fees Type', 'Total Fees', 'Single Payment Amount', 'Single Payment Paid Date',
                'EMI 1 Date', 'EMI 1 Amount',
                'EMI 2 Date', 'EMI 2 Amount',
                'EMI 3 Date', 'EMI 3 Amount',
                'EMI 4 Date', 'EMI 4 Amount',
                'EMI 5 Date', 'EMI 5 Amount',
                'EMI 6 Date', 'EMI 6 Amount',
                'Balance Amount'
            ])

            for payment in selected_payments:
                print(f"Processing payment: {payment}")

                # Fetch single payment if exists
                single_payment = SinglePayment.objects.filter(payment_info=payment).first()

                # Get all subclasses of BaseEMI
                emi_subclasses = [model for model in apps.get_models() if issubclass(model, BaseEMI)]

                # Fetch installments from all subclasses
                installments = []
                for subclass in emi_subclasses:
                    installments += list(subclass.objects.filter(payment_info=payment))

                # Sort installments by payment date to ensure the latest date is kept
                installments.sort(key=lambda x: x.date if x.date else datetime.min)

                emi_dates = ['N/A'] * 6  # Initialize 6 EMI dates (one for each EMI)
                emi_amounts = [Decimal(0)] * 6  # Initialize 6 EMI amounts (one for each EMI)

                # Dictionary to store summed amounts by installment type
                emi_amounts_dict = {}
                emi_last_date_dict = {}

                # Loop through all installments
                for installment in installments:
                    emi_type = installment.__class__.__name__  # Identify the EMI type by its class name
                    emi_index = int(emi_type[-1]) - 1  # Get the EMI number (1-6) and convert to index (0-5)

                    # Sum the amounts for the same EMI and retain the last date
                    if emi_index in emi_amounts_dict:
                        emi_amounts_dict[emi_index] += Decimal(installment.amount or 0)
                    else:
                        emi_amounts_dict[emi_index] = Decimal(installment.amount or 0)

                    # Update the last date for the EMI
                    emi_last_date_dict[emi_index] = installment.date

                # Now that all installments are processed, populate the emi_dates and emi_amounts lists
                for emi_index in range(6):
                    if emi_index in emi_last_date_dict:
                        emi_dates[emi_index] = emi_last_date_dict[emi_index].strftime('%d-%m-%Y') if emi_last_date_dict[emi_index] else 'N/A'
                    if emi_index in emi_amounts_dict:
                        emi_amounts[emi_index] = emi_amounts_dict[emi_index] if emi_amounts_dict[emi_index] > 0 else 'N/A'

                # Convert total_fees to Decimal to ensure compatibility
                try:
                    total_fees = Decimal(payment.total_fees)
                except Exception as e:
                    total_fees = Decimal(0)

                # Calculate balance amount
                total_amount_paid = sum([amt if isinstance(amt, Decimal) else Decimal(0) for amt in emi_amounts])
                balance_amount = total_fees - total_amount_paid

                # Write data row
                writer.writerow([
                    payment.registration_no,
                    payment.joining_date,
                    payment.student_name,
                    payment.course_name,
                    payment.duration,
                    payment.get_fees_type_display(),  # Get readable label of fees type
                    payment.total_fees,
                    single_payment.amount if single_payment else 'N/A',  # Single payment amount
                    single_payment.paid_date if single_payment else 'N/A',  # Single payment paid date
                    emi_dates[0], emi_amounts[0],
                    emi_dates[1], emi_amounts[1],
                    emi_dates[2], emi_amounts[2],
                    emi_dates[3], emi_amounts[3],
                    emi_dates[4], emi_amounts[4],
                    emi_dates[5], emi_amounts[5],
                    balance_amount  # Balance amount
                ])

            return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    return HttpResponse(status=400)  # Bad request if not POST or AJAX

@csrf_exempt
def export_payment_excel(request):
    if request.method == 'POST':
        try:
            ids = request.POST.get('ids', '').split(',')  # Get the ids from AJAX request

            # Fetch selected payments based on IDs
            selected_payments = PaymentInfo.objects.filter(id__in=ids)
            
            if not selected_payments:
                return JsonResponse({'error': 'No Payment available.'}, status=404)

            # Create an Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active

            # Define header row
            headers = [
                'Registration No', 'Joining Date', 'Student Name', 'Course Name', 'Course Duration', 
                'Fees Type', 'Total Fees', 'Single Payment Date', 'Single Payment Mode', 'Single Payment Amount',
                'EMI 1 Date', 'EMI 1 Amount', 'EMI 2 Date', 'EMI 2 Amount', 
                'EMI 3 Date', 'EMI 3 Amount', 'EMI 4 Date', 'EMI 4 Amount',
                'EMI 5 Date', 'EMI 5 Amount', 'EMI 6 Date', 'EMI 6 Amount', 
                'Balance Amount'
            ]
            
            # Append the header row to the sheet
            ws.append(headers)

            for payment in selected_payments:
                # Fetch single payment if exists
                single_payment = SinglePayment.objects.filter(payment_info=payment).first()

                # Get all subclasses of BaseEMI
                emi_subclasses = [model for model in apps.get_models() if issubclass(model, BaseEMI)]

                # Fetch installments from all subclasses
                installments = []
                for subclass in emi_subclasses:
                    installments += list(subclass.objects.filter(payment_info=payment))

                # Sort installments by payment date to ensure the latest date is kept
                installments.sort(key=lambda x: x.date if x.date else datetime.min)

                emi_dates = ['N/A'] * 6  # Initialize 6 EMI dates
                emi_amounts = ['N/A'] * 6  # Initialize 6 EMI amounts

                # Dictionary to store summed amounts by installment type
                emi_amounts_dict = {}
                emi_last_date_dict = {}

                # Loop through all installments
                for installment in installments:
                    emi_type = installment.__class__.__name__  # Identify the EMI type by its class name
                    emi_index = int(emi_type[-1]) - 1  # Get the EMI number (1-6) and convert to index (0-5)

                    # Sum the amounts for the same EMI and retain the last date
                    if emi_index in emi_amounts_dict:
                        emi_amounts_dict[emi_index] += Decimal(installment.amount or 0)
                    else:
                        emi_amounts_dict[emi_index] = Decimal(installment.amount or 0)

                    # Update the last date for the EMI
                    emi_last_date_dict[emi_index] = installment.date

                # Populate the emi_dates and emi_amounts lists
                for emi_index in range(6):
                    if emi_index in emi_last_date_dict:
                        emi_dates[emi_index] = emi_last_date_dict[emi_index].strftime('%d-%m-%Y') if emi_last_date_dict[emi_index] else 'N/A'
                    if emi_index in emi_amounts_dict:
                        # If amount is 0 or None, return 'N/A'
                        emi_amounts[emi_index] = emi_amounts_dict[emi_index] if emi_amounts_dict[emi_index] > 0 else 'N/A'

                # Convert total_fees to Decimal to ensure compatibility
                try:
                    total_fees = Decimal(payment.total_fees)
                except Exception as e:
                    total_fees = Decimal(0)

                # Calculate balance amount
                total_amount_paid = sum([amt if isinstance(amt, Decimal) else Decimal(0) for amt in emi_amounts])
                balance_amount = total_fees - total_amount_paid

                # Write data row
                ws.append([
                    payment.registration_no,
                    payment.joining_date,
                    payment.student_name,
                    payment.course_name,
                    payment.duration,
                    payment.get_fees_type_display(),  # Get readable label of fees type
                    payment.total_fees,
                    single_payment.date.strftime('%d-%m-%Y') if single_payment else 'N/A',  # Single payment date
                    single_payment.get_payment_mode_display() if single_payment else 'N/A',  # Single payment mode
                    single_payment.amount if single_payment and single_payment.amount else 'N/A',  # Single payment amount
                    emi_dates[0], emi_amounts[0],
                    emi_dates[1], emi_amounts[1],
                    emi_dates[2], emi_amounts[2],
                    emi_dates[3], emi_amounts[3],
                    emi_dates[4], emi_amounts[4],
                    emi_dates[5], emi_amounts[5],
                    balance_amount  # Balance amount
                ])

            # Create an in-memory file-like object to save the workbook
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            # Create the HTTP response with Excel content type and attachment header
            response = HttpResponse(output, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
            response['Content-Disposition'] = 'attachment; filename="payment_list.xlsx"'
            
            return response

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

    # Handle GET request or non-AJAX POST request here if needed
    return HttpResponse(status=400)  # Bad request if not POST or AJAX

from django.utils import timezone

@csrf_protect
@require_POST
def export_payment_pdf(request):
    ids = request.POST.get('ids', '').split(',')
    selected_payments = PaymentInfo.objects.filter(id__in=ids)

    if not selected_payments:
        return JsonResponse({'error': 'No Payment available.'}, status=404)

    attribute_list = []
    for payment in selected_payments:
        # Fetch single payment if exists
        single_payment = SinglePayment.objects.filter(payment_info=payment).first()

        # Get all subclasses of BaseEMI
        emi_subclasses = [model for model in apps.get_models() if issubclass(model, BaseEMI)]

        # Fetch installments from all subclasses
        installments = []
        for subclass in emi_subclasses:
            installments += list(subclass.objects.filter(payment_info=payment))

        # Sort installments by payment date to ensure the latest date is kept
        installments.sort(key=lambda x: x.date if x.date else timezone.datetime.min)

        emi_dates = ['N/A'] * 6  # Initialize 6 EMI dates
        emi_amounts = ['N/A'] * 6  # Initialize 6 EMI amounts

        # Dictionary to store summed amounts by installment type
        emi_amounts_dict = {}
        emi_last_date_dict = {}

        # Loop through all installments
        for installment in installments:
            emi_type = installment.__class__.__name__  # Identify the EMI type by its class name
            emi_index = int(emi_type[-1]) - 1  # Get the EMI number (1-6) and convert to index (0-5)

            # Sum the amounts for the same EMI and retain the last date
            if emi_index in emi_amounts_dict:
                emi_amounts_dict[emi_index] += Decimal(installment.amount or 0)
            else:
                emi_amounts_dict[emi_index] = Decimal(installment.amount or 0)

            # Update the last date for the EMI
            emi_last_date_dict[emi_index] = installment.date

        # Populate the emi_dates and emi_amounts lists
        for emi_index in range(6):
            if emi_index in emi_last_date_dict:
                emi_dates[emi_index] = emi_last_date_dict[emi_index].strftime('%d-%m-%Y') if emi_last_date_dict[emi_index] else 'N/A'
            if emi_index in emi_amounts_dict:
                # If amount is 0 or None, return 'N/A'
                emi_amounts[emi_index] = emi_amounts_dict[emi_index] if emi_amounts_dict[emi_index] > 0 else 'N/A'

        # Convert total_fees to Decimal to ensure compatibility
        try:
            total_fees = Decimal(payment.total_fees)
        except Exception as e:
            total_fees = Decimal(0)

        # Calculate balance amount
        total_amount_paid = sum([amt if isinstance(amt, Decimal) else Decimal(0) for amt in emi_amounts])
        balance_amount = total_fees - total_amount_paid

        # Append payment details to attribute_list
        attribute_list.append({
            'registration_no': payment.registration_no,
            'joining_date': payment.joining_date,
            'student_name': payment.student_name,
            'course_name': payment.course_name,
            'duration': payment.duration,
            'total_fees': payment.total_fees,
            'fees_type': payment.get_fees_type_display(),
            'single_payment_date': single_payment.date.strftime('%Y-%m-%d') if single_payment else '',
            'single_payment_mode': single_payment.get_payment_mode_display() if single_payment else '',
            'single_payment_amount': single_payment.amount if single_payment else 'N/A',
            'emi_dates': emi_dates,
            'emi_amounts': emi_amounts,
            'total_payment': total_amount_paid,
            'balance': balance_amount
        })

    content = {'payment_list': attribute_list}
    return renderers.render_to_pdf('payment_data_list.html', content)

class SearchPaymentResultsView(ListView):
    model = PaymentInfo
    template_name = 'search_payment_result.html'
    context_object_name = 'payments'
    paginate_by = 10

    def get_queryset(self):
        start_date = self.request.GET.get("start_date", "")
        end_date = self.request.GET.get("end_date", "")
        query = self.request.GET.get("q", "")

        object_list = PaymentInfo.objects.prefetch_related('single_payment', 'emi_1_payments', 'emi_2_payments', 'emi_3_payments', 'emi_4_payments', 'emi_5_payments', 'emi_6_payments').all()

        # Apply text search if query is provided
        if query:
            emi_filters = Q()
            for i in range(1, 7):  # Loop through EMI payments
                emi_filters |= Q(**{f'emi_{i}_payments__payment_mode__icontains': query}) | \
                               Q(**{f'emi_{i}_payments__emi__icontains': query}) | \
                               Q(**{f'emi_{i}_payments__upi_transaction_id__icontains': query}) | \
                               Q(**{f'emi_{i}_payments__upi_app_name__icontains': query}) | \
                               Q(**{f'emi_{i}_payments__refference_no__icontains': query})

            # Combine all filters
            object_list = object_list.filter(
                Q(registration_no__icontains=query) |
                Q(joining_date__icontains=query) |
                Q(student_name__icontains=query) |
                Q(course_name__icontains=query) |
                Q(duration__icontains=query) |
                Q(total_fees__icontains=query) |
                Q(fees_type__icontains=query) |
                Q(single_payment__payment_mode__icontains=query) |
                Q(single_payment__upi_transaction_id__icontains=query) |
                Q(single_payment__upi_app_name__icontains=query) |
                emi_filters
            ).distinct()

        # Apply date range filter if valid dates are provided
        if start_date and end_date:
            try:
                start_date = datetime.strptime(start_date, '%Y-%m-%d').date()
                end_date = datetime.strptime(end_date, '%Y-%m-%d').date()
                object_list = object_list.filter(joining_date__range=(start_date, end_date))
            except ValueError:
                messages.error(self.request, "Invalid date format. Please use YYYY-MM-DD.")

        # Optimize query by selecting related objects; ensure these are also correct
        object_list = object_list.select_related('single_payment')

        return object_list

    def get_context_data(self, **kwargs):
        # Get the existing context data
        context = super().get_context_data(**kwargs)

        # Get the queryset and calculate totals
        payments = context['payments']
        totals = calculate_payment_totals(payments)

        # Add the calculated totals to the context
        context.update(totals)

        return context

def new_payment_info_view(request):
    if request.method == 'POST':
        
        registration_no = request.POST.get('registration_no')
        
        # Fetch the related Enquiry object
        try:
            enrollment = Enrollment.objects.get(registration_no=registration_no)
        except Enrollment.DoesNotExist:
            context = {
                'error': 'Enrollment with the provided Registration Number does not exist.',
            }
            return render(request, 'new_payment_info.html', context)
        
        # Auto-populate fields based on the related Enquiry object
        
        payment_data = {
            'registration_no' : request.POST.get('registration_no'),
            'joining_date': enrollment.registration_date.strftime('%Y-%m-%d'),
            'student_name': enrollment.name,
            'course_name': enrollment.course_name.course_name,
            'duration': request.POST.get('duration'),
            'fees_type': request.POST.get('fees_type'),
            'total_fees': float(enrollment.total_fees_amount),  # Convert Decimal to float
            'installment_amount': float(enrollment.installment_amount),  # Convert Decimal to float
            'montly_payment_type': request.POST.get('montly_payment_type'),
        }
        
        print("Payment Data : ", payment_data)
        

        # Validate and process the form data
        if not (payment_data.get('registration_no') and payment_data.get('student_name') and payment_data.get('course_name') and payment_data.get('duration') and payment_data.get('total_fees') and payment_data.get('fees_type') and payment_data.get('joining_date')):
            messages.error(request, "All fields are required.")
            return render(request, 'new_payment_info.html')

        try:
            token = Token.objects.get(user=request.user)
        except Token.DoesNotExist:
            context = {
                'error': 'Authentication Token not Found',
                **payment_data,
            }
            return render(request, 'new_payment_info.html', context)

        api_url = 'http://127.0.0.1:8000/api/payment_info/'
        
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json',
        }

        try:
            response = requests.post(api_url, json=payment_data, headers=headers)
            response_data = response.json()
            
            print("Response Data : ",response_data)

        except requests.exceptions.RequestException:
            context = {
                'error': 'An Error Occurred While Creating an Enrollment',
                **payment_data,
            }
            return render(request, 'new_payment_info.html', context)

        if response.status_code in [200, 201]:
            messages.success(request, 'Created Successfully')
            if payment_data.get('fees_type') == 'Regular':
                return redirect('single_payment')  # Redirect to a success page or another view
            else:
                return redirect('new_installment_info')
        else:
            error_message = response_data.get('error', 'An Error Occurred During Creation.')
            errors = response_data
            context = {
                'error': error_message,
                'errors': errors,
                **payment_data,
            }
        return render(request, 'new_payment_info.html', context)

    return render(request, 'new_payment_info.html')

def single_payment_view(request):
    if request.method == 'POST':
        payment_mode = request.POST.get('payment_mode')
        payment_data = {    
            'payment_info': request.POST.get('payment_info'),
            'date': request.POST.get('date'),
            'payment_mode': payment_mode,
            'amount': request.POST.get('amount'),
        }

        # Handle conditional fields based on payment_mode
        if payment_mode == 'UPI':
            payment_data.update({
                'upi_transaction_id': request.POST.get('upi_transaction_id'),
                'upi_app_name': request.POST.get('upi_app_name'),
            })
        elif payment_mode == 'Bank Transfer':
            payment_data.update({
                'refference_no': request.POST.get('refference_no'),
            })

        # Validate required fields based on payment_mode
        if not all(payment_data.values()):
            messages.error(request, "All fields are required.")
            return render(request, 'single_payment.html', {'payment_data': payment_data})
        
        # Authentication token (if needed) and API call logic
        try:
            token = Token.objects.get(user=request.user)
        except Token.DoesNotExist:
            return render(request, 'single_payment.html', {'error': 'Authentication Token not Found', 'payment_data': payment_data})

        api_url = 'http://127.0.0.1:8000/api/single_payment/'
        
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json',
        }

        try:
            response = requests.post(api_url, json=payment_data, headers=headers)
            response_data = response.json()
            
            print("Response Data : ",response_data)

        except requests.exceptions.RequestException:
            context = {
                'error': 'An Error Occurred While Creating an Enrollment',
                **payment_data,
            }
            return render(request, 'single_payment.html', context)

        if response.status_code in [200, 201]:
            messages.success(request, 'Created Successfully')
            return redirect('manage_payments')  # Redirect to a success page or another view
        else:
            error_message = response_data.get('error', 'An Error Occurred During Creation.')
            errors = response_data
            context = {
                'error': error_message,
                'errors': errors,
                **payment_data,
            }
        return render(request, 'single_payment.html', context)
    # For GET requests or initial form rendering
    payment_info = PaymentInfo.objects.last()
    context = {
        'payment_info': payment_info
    }
    return render(request, 'single_payment.html', context)


def single_payment_update_view(request, id):
    if request.method == 'POST':
        payment_mode = request.POST.get('payment_mode')
        payment_data = {    
            'payment_info': request.POST.get('payment_info'),
            'date': request.POST.get('date'),
            'payment_mode': payment_mode,
            'amount': request.POST.get('amount'),
        }

        # Handle conditional fields based on payment_mode
        if payment_mode == 'UPI':
            payment_data.update({
                'upi_transaction_id': request.POST.get('upi_transaction_id'),
                'upi_app_name': request.POST.get('upi_app_name'),
            })
        elif payment_mode == 'Bank Transfer':
            payment_data.update({
                'refference_no': request.POST.get('refference_no'),
            })

        # Validate required fields based on payment_mode
        if not all(payment_data.values()):
            messages.error(request, "All fields are required.")
            return render(request, 'single_payment.html', {'payment_data': payment_data})
        
        # Authentication token (if needed) and API call logic
        try:
            token = Token.objects.get(user=request.user)
        except Token.DoesNotExist:
            return render(request, 'single_payment.html', {'error': 'Authentication Token not Found', 'payment_data': payment_data})

        api_url = 'http://127.0.0.1:8000/api/single_payment/'
        headers = {
            'Authorization': f'Token {token.key}',
            'Content-Type': 'application/json',
        }

        try:
            response = requests.post(api_url, json=payment_data, headers=headers)
            response_data = response.json()
        except requests.exceptions.RequestException:
            return render(request, 'single_payment.html', {'error': 'An Error Occurred While Processing Payment', 'payment_data': payment_data})

        if response.status_code in [200, 201]:
            messages.success(request, 'Payment Processed Successfully')
            return redirect('manage_payments')
        else:
            return render(request, 'single_payment.html', {'error': response_data.get('error', 'An error occurred.'), 'payment_data': payment_data})

    # Fetch payments to display in table
    payments = SinglePayment.objects.filter(is_active=True, is_deleted=False)
    return render(request, 'single_payment.html', {'payments': payments})

from .config import EMI_MODELS  # Import the dictionary mapping

def new_installment_view(request):
    print("New Installment!!!!")
    
    payment_info = PaymentInfo.objects.last()
    if not payment_info:
        messages.error(request, "Payment information not found.")
        return redirect('new_installment_info')

    # Calculate the total paid amount by summing all EMI amounts
    total_paid_amount = sum(emi.amount for emi in EMI_MODELS['EMI_1'].objects.filter(payment_info=payment_info))
    total_remaining_fees = payment_info.total_fees - total_paid_amount
    print(f"Total Remaining Fees: {total_remaining_fees}")

    if request.method == 'POST':
        try:
            next_emi = get_next_emi(payment_info, 'EMI_1')
            print(f"EMI : {next_emi}")
        except ValueError as e:
            messages.error(request, str(e))
            return redirect('new_installment_info')

        if not next_emi:
            messages.error(request, "All EMIs are completed.")
            return redirect('new_installment_info')

        payment_mode = request.POST.get('payment_mode')
        payment_amount = Decimal(request.POST.get('amount', 0))
        
        # Condition 1: Ensure payment amount is not greater than remaining fees
        if payment_amount <= 0:
            messages.error(request, "Invalid payment amount.")
            return redirect('new_installment_info')
        if payment_amount > total_remaining_fees:
            messages.error(request, f"Entered amount exceeds the remaining total fees of {total_remaining_fees}.")
            return redirect('new_installment_info')

        remaining_amount = payment_amount
        print(f"Remaining Amount : {remaining_amount}")

        registration_no = request.POST.get('registration_no')
        date = request.POST.get('date')
        print(f"Registration No: {registration_no}, Date: {date}")

        # Process EMI_1 explicitly
        emi_model = EMI_MODELS.get('EMI_1')
        if emi_model:
            next_emi_amount = payment_info.installment_amount
            print(f"Processing full payment for EMI_1, Installment : {next_emi_amount}")
            
            status = "Pending" if remaining_amount < next_emi_amount else "Paid"
            print(f"Status : {status}")
            
            paid_emi_instance = emi_model(
                payment_info=payment_info,
                registration_no=registration_no,
                date=date,
                payment_mode=payment_mode,
                emi='EMI_1',
                amount=min(next_emi_amount, remaining_amount),
                status=status
            )
            
            # Set additional fields based on payment mode
            if payment_mode == 'Bank Transfer':
                paid_emi_instance.refference_no = request.POST.get('refference_no')
            elif payment_mode == 'UPI':
                paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
                paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

            paid_emi_instance.save()  # Save after setting all fields
            remaining_amount -= next_emi_amount
            print(f"Processed Full Payment for EMI_1: Status: {paid_emi_instance.status}, Remaining Amount: {remaining_amount}")

        # Process subsequent EMIs
        while remaining_amount > 0:
            next_emi = get_next_emi(payment_info, next_emi.emi)  # Get the next EMI for processing
            if not next_emi:
                break  # No more EMIs to process

            emi_model = EMI_MODELS.get(next_emi.emi)
            if not emi_model:
                messages.error(request, f"EMI model not found for {next_emi.emi}.")
                return redirect('new_installment_info')

            next_emi_amount = payment_info.installment_amount
            print(f"Next EMI Amount : {next_emi_amount}, Remaining Amount : {remaining_amount}")

            if remaining_amount >= next_emi_amount:
                # Full payment for the current EMI
                print(f"Processing full payment for {next_emi.emi}")
                paid_emi_instance = emi_model(
                    payment_info=payment_info,
                    registration_no=registration_no,
                    date=date,
                    payment_mode=payment_mode,
                    emi=next_emi.emi,
                    amount=next_emi_amount,
                    status="Paid"
                )
                
                # Set additional fields based on payment mode
                if payment_mode == 'Bank Transfer':
                    paid_emi_instance.refference_no = request.POST.get('refference_no')
                elif payment_mode == 'UPI':
                    paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
                    paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

                paid_emi_instance.save()
                remaining_amount -= next_emi_amount
                print(f"Processed Full Payment for {next_emi.emi}: Status: {paid_emi_instance.status}, Remaining Amount: {remaining_amount}")

            else:
                # Partial payment for the current EMI
                print(f"Processing partial payment for {next_emi.emi}")
                partial_payment_instance = emi_model(
                    payment_info=payment_info,
                    registration_no=registration_no,
                    date=date,
                    payment_mode=payment_mode,
                    emi=next_emi.emi,
                    amount=remaining_amount,
                    status="Pending"
                )
                
                # Set additional fields based on payment mode
                if payment_mode == 'Bank Transfer':
                    partial_payment_instance.refference_no = request.POST.get('refference_no')
                elif payment_mode == 'UPI':
                    partial_payment_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
                    partial_payment_instance.upi_app_name = request.POST.get('upi_app_name')

                partial_payment_instance.save()
                next_emi.amount -= remaining_amount
                remaining_amount = 0  # All remaining amount is paid
                print(f"Processed Partial Payment for {next_emi.emi}: Status: {partial_payment_instance.status}")

        messages.success(request, 'Installment created successfully.')
        return redirect('new_manage_payments')

    else:
        context = {
            'next_emi': "EMI_1",
            'payment_info': payment_info,
        }
        return render(request, 'new_installment_info.html', context)

def new_manage_payment_info_view(request):
    try:
        token = Token.objects.get(user=request.user)
    except Token.DoesNotExist:
        context = {'error': 'Authentication token not found'}
        return render(request, 'manage_payment_info.html', context)

    api_url = 'http://127.0.0.1:8000/api/payment_info/'
    headers = {
        'Authorization': f'Token {token.key}',
        'Content-Type': 'application/json'
    }

    try:
        response = requests.get(api_url, headers=headers)
        response.raise_for_status()
        response_data = response.json()
        
    except requests.exceptions.RequestException as err:
        context = {
            'error': f'Request error occurred: {err}',
            'response_data': []
        }
        return render(request, 'manage_payment_info.html', context)

    per_page = request.GET.get('per_page', '10')
    paginator = Paginator(response_data, per_page)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)

    emi_data = []  # List to store EMI details
    payment_totals = {}  # Dictionary to hold total amounts for each payment_info_id
    total_emi_sums = {f'emi_{i}': 0 for i in range(1, 7)}  # Dictionary to hold totals for emi_1 to emi_6
    
    for payment in response_data:
        for i in range(1, 7):
            emi_key = f'emi_{i}_payments'
            emi_payments = payment.get(emi_key, [])

            if not isinstance(emi_payments, list):
                # print(f"Expected a list for {emi_key}, but got: {emi_payments}")
                continue

            total_amount = 0  # Initialize total_amount for this emi_key
            dates = []  # Initialize dates list for this emi_key
            payment_info_ids = set()  # Use a set to avoid duplicates

            for emi in emi_payments:
                if isinstance(emi, dict):
                    payment_info_id = emi.get('payment_info', None)

                    if payment_info_id is not None:
                        payment_info_ids.add(payment_info_id)  # Collect unique payment IDs
                        total_amount += float(emi.get('amount', 0))
                        if emi['date']:
                            formatted_date = datetime.strptime(emi['date'], '%Y-%m-%d').strftime('%d-%m-%Y')
                            dates.append(formatted_date)
                        else:
                            print("Date is None for EMI:", emi)
            # Append to emi_data after processing all EMI payments for the current payment
            emi_data.append({
                'emi_key': emi_key,
                'total_amount': total_amount,
                'dates': dates,
                'payment_ids': list(payment_info_ids),  # Convert set to list for context
            })

            # Sum the total for each emi payments
            total_emi_sums[f'emi_{i}'] += total_amount
            
    # Initialize a dictionary to hold total amounts per payment_info_id
    payment_totals = {}
    # First pass to collect total amounts
    for emi in emi_data:
        total_amount = emi['total_amount']
        for payment_id in emi['payment_ids']:
            if payment_id not in payment_totals:
                payment_totals[payment_id] = 0
            payment_totals[payment_id] += total_amount  # Aggregate the total for this payment ID

    # Now calculate the final amounts using the payment_totals
    final_amounts = {}
    total_final_amount_sum = 0
    for payment_id in payment_totals:
        total_amount = payment_totals[payment_id]
        # print(f"Total amount for payment_info_id {payment_id}: {total_amount}")

        # Get total fees from PaymentInfo based on payment_id
        try:
            payment_info = PaymentInfo.objects.get(id=payment_id)
            total_fees = Decimal(payment_info.total_fees)
        except PaymentInfo.DoesNotExist:
            total_fees = Decimal(0)
            # print(f"No fees found for payment_info_id {payment_id}, setting total fees to 0")

        # print(f"Total Fees for payment_info_id {payment_id}: {total_fees}")

        # Calculate final amount as total_fees - total_amount
        final_amount = total_fees - Decimal(total_amount)
        final_amounts[payment_id] = final_amount
        # print(f"Final Amount for payment_info_id {payment_id}: {final_amount}")

        # Accumulate the sum of final amounts
        total_final_amount_sum += final_amount
    
    # Calculate total fees sum from response_data
    total_fees_sum = sum(
        Decimal(entry.get('total_fees', 0)) if entry.get('total_fees') else 0
        for entry in response_data
    )
    
    context = {
        'page_obj': page_obj,
        'per_page': per_page,
        'payment_data': response_data,
        'emi_data': emi_data,  # Include emi_data in context
        'final_amounts': final_amounts,  # Include final amounts in context
        'total_fees_sum': total_fees_sum,  # Include final amounts in context
        'total_emi_sums': total_emi_sums,
        'total_final_amount_sum': total_final_amount_sum,
        'emi_range': range(1, 7),
    }

    return render(request, 'manage_payment_info.html', context)

# def new_installment_update_view(request, id):
#     payment_info = get_object_or_404(PaymentInfo, id=id)
#     print(f"Payment Info: {payment_info}")  # Debug

#     total_fees = payment_info.total_fees
    
#     # Calculate total paid amount by summing over all EMI models
#     total_paid_amount = 0
#     total_pending_amount = 0  # To track pending amount
    
#     # Calculate total paid amount by summing over all EMI models
#     for emi_model in EMI_MODELS.values():
#         # Sum amounts for paid EMIs
#         total_paid_amount += sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info, status='Paid'))

#         # Sum amounts for pending EMIs
#         total_pending_amount += sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info, status='Pending'))

#     remaining_balance = total_fees - (total_paid_amount + total_pending_amount) # Calculate remaining balance

#     print(f"remaining balance : {remaining_balance} , Total Paid Amount : {total_paid_amount} , total_pending_amount : {total_pending_amount}")
    
    
#     if request.method == 'POST':
#         payment_amount = Decimal(request.POST.get('amount', 0))
#         registration_no = request.POST.get('registration_no')
#         payment_mode = request.POST.get('payment_mode') 
#         input_date = request.POST.get('date')
        
#         parsed_date = datetime.strptime(input_date, "%d-%m-%Y")  # Adjust if input format changes
#         date = parsed_date.strftime("%Y-%m-%d")

#         # Validation for payment amount
#         if payment_amount <= 0:
#             messages.error(request, "Invalid payment amount.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         # Ensure payment amount does not exceed remaining balance
#         if payment_amount > remaining_balance:
#             messages.error(request, f"Payment amount exceeds the remaining balance. Remaining balance: {remaining_balance}")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         if not date:
#             messages.error(request, "Date is required.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         try:
#             parsed_date = datetime.strptime(date, '%Y-%m-%d')
#         except ValueError:
#             messages.error(request, "Invalid date format.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         last_pending_emi = get_last_pending_emi(payment_info)
#         emi_type = None
#         remaining_amount = 0

#         if isinstance(last_pending_emi, str):
#             emi_type = last_pending_emi
#         elif hasattr(last_pending_emi, 'emi'):
#             emi_type = last_pending_emi.emi
#             remaining_amount = last_pending_emi.amount

#         if not emi_type:
#             messages.error(request, "No EMIs available for processing.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         while payment_amount > 0:
#             next_emi = get_next_emi(payment_info, emi_type)
            
#             print(f"Next EMI : {next_emi}")
            
#             if next_emi is None:
#                 break

#             emi_model = EMI_MODELS.get(next_emi.emi)
            
#             print(f"EMI Model : {emi_model}")
            
#             if not emi_model:
#                 messages.error(request, f"EMI model not found for {next_emi.emi}.")
#                 return redirect('new_installment_update_info', id=payment_info.id)

#             next_emi_amount = next_emi.amount
#             print(f"Pending Amount : {next_emi_amount}")
            
#             installment = payment_info.installment_amount
#             print(f"Installment : {installment}")
            
#             remaining_amount = installment - next_emi_amount
#             print(f"Processing EMI: {next_emi.emi}, Amount: {next_emi_amount}, Remaining Payment Amount: {remaining_amount}")  # Debug
            
#             if payment_amount >= next_emi_amount:
#                 paid_emi_instance = emi_model(
#                     payment_info=payment_info,
#                     registration_no=registration_no,
#                     date=parsed_date,
#                     payment_mode=payment_mode,
#                     emi=next_emi.emi,
#                     amount= remaining_amount if get_last_emi_status(payment_info, next_emi.emi) == "Pending" else payment_info.installment_amount,
#                     status="Paid"
#                 )
#                 paid_emi_instance.save()
#                 payment_amount -= paid_emi_instance.amount
#                 print(f"Payment Amount After Saving : {payment_amount}, Payment paid Amount : {paid_emi_instance.amount}")
                
#                 # next_emi.amount = 0
#             else:
#                 paid_emi_instance = emi_model(
#                     payment_info=payment_info,
#                     registration_no=registration_no,
#                     date=parsed_date,
#                     payment_mode=payment_mode,
#                     emi=next_emi.emi,
#                     amount=payment_amount,
#                     status="Pending"
#                 )
#                 paid_emi_instance.save()

#                 # next_emi.amount -= payment_amount
#                 payment_amount = 0

#         # Add additional fields based on payment mode
#         if payment_mode == 'Bank Transfer':
#             paid_emi_instance.refference_no = request.POST.get('refference_no')
#         elif payment_mode == 'UPI':
#             paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
#             paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

#         try:
#             paid_emi_instance.save()
#             print("Paid EMI instance saved successfully.")  # Debug
#         except Exception as e:
#             print(f"Error saving instance: {e}")  # Debug
#             messages.error(request, "Error saving the payment information.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         messages.success(request, 'Payment processed successfully.')
#         return redirect('new_manage_payments')

#     # For GET requests or initial form rendering
#     last_pending_emi = get_last_pending_emi(payment_info)

#     if isinstance(last_pending_emi, str):
#         emi_type = last_pending_emi
#     elif hasattr(last_pending_emi, 'emi'):
#         emi_type = last_pending_emi.emi
#     else:
#         messages.error(request, "No EMIs available for processing.")
#         return redirect('new_installment_update_info', id=payment_info.id)

#     # Retrieve installments to display in the table
#     installments = []
#     for emi_model in EMI_MODELS.values():
#         installments.extend(emi_model.objects.filter(payment_info=payment_info))
    
#     context = {
#         'next_emi': emi_type,
#         'payment_info': payment_info,
#         'remaining_balance': remaining_balance,  # Send remaining balance to template
#         'installments': installments  # Assuming this is how you fetch installments
#     }

#     return render(request, 'new_installment_info.html', context)


def get_last_emi_status(payment_info, emi_value):
    # Determine the correct EMI model based on emi_value
    emi_model = EMI_MODELS.get(emi_value)

    if not emi_model:
        return None  # Return None if the model is not found

    # Fetch the last EMI based on emi_value, ordered by date
    last_emi = emi_model.objects.filter(payment_info=payment_info).order_by('-date').first()
    return last_emi.status if last_emi else None



def get_next_emi(payment_info, emi_type):
    EMI_Model = EMI_MODELS.get(emi_type)
    print(f"Retrieved EMI Model for {emi_type}: {EMI_Model}")
    if not EMI_Model:
        raise ValueError(f"Invalid EMI type: {emi_type}")

    installments = EMI_Model.objects.filter(payment_info=payment_info)
    print(f"Existing installments: {installments}")

    if not installments:
        # Create first EMI if none exists
        next_emi = EMI_Model(emi=emi_type, amount=payment_info.installment_amount)
        next_emi.payment_info = payment_info
        # next_emi.save()
        print(f"Created first EMI: {next_emi}")
        return next_emi

    last_installment = max(installments, key=lambda x: x.created_at)
    print(f"Last installment: {last_installment}")

    if last_installment.status == 'Pending':
        return last_installment

    next_emi_number = int(last_installment.emi.split('_')[1]) + 1
    print(f"Next EMI number: {next_emi_number}")

    if next_emi_number > 6:  # Assuming 6 installments
        print("No more EMIs available.")
        return None

    next_emi = EMI_Model(emi=f"{last_installment.emi.split('_')[0]}_{next_emi_number}", amount=payment_info.installment_amount)
    next_emi.payment_info = payment_info
    # next_emi.save()
    print(f"Created next EMI: {next_emi}")
    return next_emi

def get_last_pending_emi(payment_info):
    all_installments = []

    # Loop through all EMI types
    for emi_type in ['EMI_1', 'EMI_2', 'EMI_3', 'EMI_4', 'EMI_5', 'EMI_6']:
        EMI_Model = EMI_MODELS.get(emi_type)
        if EMI_Model:
            # Collect both pending and paid installments
            installments = EMI_Model.objects.filter(payment_info=payment_info).order_by('-created_at')
            print(f"Processing {emi_type}: Found {installments.count()} installments.")
            
            if installments.exists():
                recent_installment = installments.first()
                print(f"Most recent installment for {emi_type}: {recent_installment}")
                all_installments.append(recent_installment)  # Get the most recent installment

    # Determine the last EMI if available
    if all_installments:
        last_emi = max(all_installments, key=lambda x: x.created_at)  # Most recent by created_at
        print(f"Last EMI (Pending or Paid): {last_emi}")
        return last_emi

    print("No EMIs found.")
    return None

class EMI_1ListCreateView(generics.ListCreateAPIView):
    queryset = EMI_1.objects.all().order_by('-id')
    serializer_class = EMI_1_Serializer
    permission_classes = [IsAuthenticated]

class EMI_1DetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EMI_1.objects.all()
    serializer_class = EMI_1_Serializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EMI_2ListCreateView(generics.ListCreateAPIView):
    queryset = EMI_2.objects.all().order_by('-id')
    serializer_class = EMI_2_Serializer
    permission_classes = [IsAuthenticated]

class EMI_2DetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EMI_2.objects.all()
    serializer_class = EMI_2_Serializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EMI_3ListCreateView(generics.ListCreateAPIView):
    queryset = EMI_3.objects.all().order_by('-id')
    serializer_class = EMI_3_Serializer
    permission_classes = [IsAuthenticated]

class EMI_3DetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EMI_3.objects.all()
    serializer_class = EMI_3_Serializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EMI_4ListCreateView(generics.ListCreateAPIView):
    queryset = EMI_4.objects.all().order_by('-id')
    serializer_class = EMI_4_Serializer
    permission_classes = [IsAuthenticated]

class EMI_4DetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EMI_4.objects.all()
    serializer_class = EMI_4_Serializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EMI_5ListCreateView(generics.ListCreateAPIView):
    queryset = EMI_5.objects.all().order_by('-id')
    serializer_class = EMI_5_Serializer
    permission_classes = [IsAuthenticated]

class EMI_5DetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EMI_5.objects.all()
    serializer_class = EMI_5_Serializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'

class EMI_6ListCreateView(generics.ListCreateAPIView):
    queryset = EMI_6.objects.all().order_by('-id')
    serializer_class = EMI_6_Serializer
    permission_classes = [IsAuthenticated]

class EMI_6DetailView(generics.RetrieveUpdateDestroyAPIView):
    queryset = EMI_6.objects.all()
    serializer_class = EMI_6_Serializer
    permission_classes = [IsAuthenticated]
    lookup_field = 'pk'


# def payment_view(request):
#     if request.method == 'POST':
#         registration_no = request.POST.get('registration_no')
        
#         # Debug print
#         print(f"Received registration_no: {registration_no}")

#         # Fetch the related Enrollment object
#         try:
#             enrollment = Enrollment.objects.get(registration_no=registration_no)
#             print(f"Enrollment found: {enrollment}")
#         except Enrollment.DoesNotExist:
#             context = {
#                 'error': 'Enrollment with the provided Registration Number does not exist.',
#             }
#             return render(request, 'new_payment_info.html', context)

#         # Validate POST data
#         duration = request.POST.get('duration')
#         fees_type = request.POST.get('fees_type')
#         monthly_payment_type = request.POST.get('montly_payment_type')
#         payment_mode = request.POST.get('payment_mode')
#         amount = request.POST.get('amount', 0)
        
#         print(f"duration : {duration}, fees_type : {fees_type}, monthly Payment type : {monthly_payment_type}, payment mode : {payment_mode}, amount : {amount}")

#         if not all([duration, fees_type, monthly_payment_type, payment_mode, amount]):
#             messages.error(request, "All fields are required.")
#             return redirect('new_payment_info')

#         # Prepare PaymentInfo data
#         payment_info_data = {
#             'registration_no': registration_no,
#             'joining_date': enrollment.registration_date,
#             'student_name': enrollment.name,
#             'course_name': enrollment.course_name.course_name,
#             'duration': duration,
#             'fees_type': fees_type,
#             'total_fees': Decimal(enrollment.total_fees_amount),
#             'installment_amount': Decimal(enrollment.installment_amount),
#         }

#         # Create a PaymentInfo instance
#         payment_info = PaymentInfo(**payment_info_data)
#         payment_info.save()
#         print(f"PaymentInfo created: {payment_info}")

#         # Prepare SinglePayment data
#         if fees_type == 'Regular':
        
#             single_payment_data = {
#                 'payment_info': payment_info,
#                 'date': request.POST.get('date'),
#                 'payment_mode': request.POST.get('payment_mode'),
#                 'amount': request.POST.get('amount'),
#             }

#             # Handle UPI and Bank Transfer specific fields
#             if payment_mode == 'Bank Transfer':
#                 single_payment_data['reference_no'] = request.POST.get('reference_no')
#             elif payment_mode == 'UPI':
#                 single_payment_data['upi_transaction_id'] = request.POST.get('upi_transaction_id')
#                 single_payment_data['upi_app_name'] = request.POST.get('upi_app_name')

#             # Create a SinglePayment instance
#             single_payment = SinglePayment(**single_payment_data)
#             single_payment.save()
#             print(f"SinglePayment created: {single_payment}")

#         # Calculate total paid amount and remaining fees
#         total_paid_amount = sum(emi.amount for emi in EMI_MODELS['EMI_1'].objects.filter(payment_info=payment_info))
#         total_remaining_fees = payment_info.total_fees - total_paid_amount
#         print(f"Total paid amount: {total_paid_amount}, Total remaining fees: {total_remaining_fees}")

#         # Validate payment amount
#         payment_amount = Decimal(amount)
#         if payment_amount <= 0:
#             messages.error(request, "Invalid payment amount.")
#             return redirect('new_payment_info')
#         if payment_amount > total_remaining_fees:
#             messages.error(request, f"Entered amount exceeds the remaining total fees of {total_remaining_fees}.")
#             return redirect('new_payment_info')

#         remaining_amount = payment_amount
#         date = request.POST.get('date')

#         # Process EMI payments
#         try:
#             next_emi = get_next_emi(payment_info, 'EMI_1')
#             print(f"Next EMI: {next_emi}")
#         except ValueError as e:
#             messages.error(request, str(e))
#             return redirect('new_installment_info')

#         # Process EMI_1 explicitly
#         emi_model = EMI_MODELS.get('EMI_1')
#         if emi_model:
#             next_emi_amount = payment_info.installment_amount
#             status = "Pending" if remaining_amount < next_emi_amount else "Paid"

#             paid_emi_instance = emi_model(
#                 payment_info=payment_info,
#                 registration_no=registration_no,
#                 date=date,
#                 payment_mode=payment_mode,
#                 emi='EMI_1',
#                 amount=min(next_emi_amount, remaining_amount),
#                 status=status
#             )

#             # Set additional fields based on payment mode
#             if payment_mode == 'Bank Transfer':
#                 paid_emi_instance.reference_no = request.POST.get('reference_no')
#             elif payment_mode == 'UPI':
#                 paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
#                 paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

#             paid_emi_instance.save()
#             remaining_amount -= next_emi_amount
#             print(f"Processed EMI_1: Amount paid: {min(next_emi_amount, remaining_amount)}, Remaining Amount: {remaining_amount}")

#         # Process subsequent EMIs
#         while remaining_amount > 0:
#             next_emi = get_next_emi(payment_info, next_emi.emi)
#             if not next_emi:
#                 print("No more EMIs to process.")
#                 break

#             emi_model = EMI_MODELS.get(next_emi.emi)
#             if not emi_model:
#                 messages.error(request, f"EMI model not found for {next_emi.emi}.")
#                 return redirect('new_installment_info')

#             next_emi_amount = payment_info.installment_amount
#             print(f"Next EMI Amount: {next_emi_amount}, Remaining Amount: {remaining_amount}")

#             if remaining_amount >= next_emi_amount:
#                 # Full payment for the current EMI
#                 paid_emi_instance = emi_model(
#                     payment_info=payment_info,
#                     registration_no=registration_no,
#                     date=date,
#                     payment_mode=payment_mode,
#                     emi=next_emi.emi,
#                     amount=next_emi_amount,
#                     status="Paid"
#                 )

#                 # Set additional fields based on payment mode
#                 if payment_mode == 'Bank Transfer':
#                     paid_emi_instance.reference_no = request.POST.get('reference_no')
#                 elif payment_mode == 'UPI':
#                     paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
#                     paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

#                 paid_emi_instance.save()
#                 remaining_amount -= next_emi_amount
#                 print(f"Processed Full Payment for {next_emi.emi}: Remaining Amount: {remaining_amount}")

#             else:
#                 # Partial payment for the current EMI
#                 partial_payment_instance = emi_model(
#                     payment_info=payment_info,
#                     registration_no=registration_no,
#                     date=date,
#                     payment_mode=payment_mode,
#                     emi=next_emi.emi,
#                     amount=remaining_amount,
#                     status="Pending"
#                 )

#                 # Set additional fields based on payment mode
#                 if payment_mode == 'Bank Transfer':
#                     partial_payment_instance.reference_no = request.POST.get('reference_no')
#                 elif payment_mode == 'UPI':
#                     partial_payment_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
#                     partial_payment_instance.upi_app_name = request.POST.get('upi_app_name')

#                 partial_payment_instance.save()
#                 next_emi.amount -= remaining_amount
#                 remaining_amount = 0  # All remaining amount is paid
#                 print(f"Processed Partial Payment for {next_emi.emi}: Remaining Amount: {remaining_amount}")

#         messages.success(request, 'Payment processed successfully!')
#         return redirect('payment')
    
#     context = {
#         'next_emi': "EMI_1",
#     }

#     return render(request, 'payment_info.html', context)

def payment_view(request):
    if request.method == 'POST':
        registration_no = request.POST.get('registration_no')
        
        # Debug print
        print(f"Received registration_no: {registration_no}")

        # Fetch the related Enrollment object
        try:
            enrollment = Enrollment.objects.get(registration_no=registration_no)
            print(f"Enrollment found: {enrollment}")
        except Enrollment.DoesNotExist:
            context = {
                'error': 'Enrollment with the provided Registration Number does not exist.',
            }
            return render(request, 'payment_info.html', context)

        # Validate POST data
        duration = request.POST.get('duration')
        fees_type = request.POST.get('fees_type')
        monthly_payment_type = request.POST.get('montly_payment_type')
        payment_mode = request.POST.get('payment_mode')
        amount = request.POST.get('amount', 0)
        
        print(f"duration : {duration}, fees_type : {fees_type}, monthly Payment type : {monthly_payment_type}, payment mode : {payment_mode}, amount : {amount}")

        if not all([duration, fees_type, monthly_payment_type, payment_mode, amount]):
            messages.error(request, "All fields are required.")
            return redirect('payment')

        # Prepare PaymentInfo data
        payment_info_data = {
            'registration_no': registration_no,
            'joining_date': enrollment.registration_date.strftime('%Y-%m-%d') if enrollment.registration_date else '',
            'student_name': enrollment.name,
            'course_name': enrollment.course_name.course_name,
            'duration': duration,
            'fees_type': fees_type,
            'total_fees': Decimal(enrollment.total_fees_amount),
            'installment_amount': Decimal(enrollment.installment_amount),
        }

        # Create a PaymentInfo instance
        payment_info = PaymentInfo(**payment_info_data)
        payment_info.save()
        print(f"PaymentInfo created: {payment_info}")

        input_date = request.POST.get('date')
        
        
        parsed_date = datetime.strptime(input_date, "%d-%m-%Y")  # Adjust if input format changes
        single_payment_date = parsed_date.strftime("%Y-%m-%d")
        
        
        # Prepare SinglePayment data
        if fees_type == 'Regular':
        
            single_payment_data = {
                'payment_info': payment_info,
                'date': single_payment_date,
                'payment_mode': request.POST.get('payment_mode'),
                'amount': request.POST.get('amount'),
            }

            # Handle UPI and Bank Transfer specific fields
            if payment_mode == 'Bank Transfer':
                single_payment_data['reference_no'] = request.POST.get('reference_no')
            elif payment_mode == 'UPI':
                single_payment_data['upi_transaction_id'] = request.POST.get('upi_transaction_id')
                single_payment_data['upi_app_name'] = request.POST.get('upi_app_name')

            # Create a SinglePayment instance
            single_payment = SinglePayment(**single_payment_data)
            single_payment.save()
            print(f"SinglePayment created: {single_payment}")

        elif fees_type == 'Installment':
            # Calculate total paid amount and remaining fees
            total_paid_amount = sum(emi.amount for emi in EMI_MODELS['EMI_1'].objects.filter(payment_info=payment_info))
            total_remaining_fees = payment_info.total_fees - total_paid_amount
            print(f"Total paid amount: {total_paid_amount}, Total remaining fees: {total_remaining_fees}")

            # Validate payment amount
            payment_amount = Decimal(amount)
            if payment_amount <= 0:
                messages.error(request, "Invalid payment amount.")
                return redirect('payment')
            if payment_amount > total_remaining_fees:
                messages.error(request, f"Entered amount exceeds the remaining total fees of {total_remaining_fees}.")
                return redirect('payment')

            remaining_amount = payment_amount
            input_date = request.POST.get('date')
            parsed_date = datetime.strptime(input_date, "%d-%m-%Y")  # Adjust if input format changes
            date = parsed_date.strftime("%Y-%m-%d")
            # Process EMI payments
            try:
                next_emi = get_next_emi(payment_info, 'EMI_1')
                print(f"Next EMI: {next_emi}")
            except ValueError as e:
                messages.error(request, str(e))
                return redirect('new_installment_info')

            # Process EMI_1 explicitly
            emi_model = EMI_MODELS.get('EMI_1')
            if emi_model:
                next_emi_amount = payment_info.installment_amount
                status = "Pending" if remaining_amount < next_emi_amount else "Paid"

                paid_emi_instance = emi_model(
                    payment_info=payment_info,
                    registration_no=registration_no,
                    date=date,
                    payment_mode=payment_mode,
                    emi='EMI_1',
                    amount=min(next_emi_amount, remaining_amount),
                    monthly_payment_type = monthly_payment_type,
                    status=status
                )

                # Set additional fields based on payment mode
                if payment_mode == 'Bank Transfer':
                    paid_emi_instance.reference_no = request.POST.get('reference_no')
                elif payment_mode == 'UPI':
                    paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
                    paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

                paid_emi_instance.save()
                remaining_amount -= next_emi_amount
                print(f"Processed EMI_1: Amount paid: {min(next_emi_amount, remaining_amount)}, Remaining Amount: {remaining_amount}")

            # Process subsequent EMIs
            while remaining_amount > 0:
                next_emi = get_next_emi(payment_info, next_emi.emi)
                if not next_emi:
                    print("No more EMIs to process.")
                    break

                emi_model = EMI_MODELS.get(next_emi.emi)
                if not emi_model:
                    messages.error(request, f"EMI model not found for {next_emi.emi}.")
                    return redirect('new_installment_info')

                next_emi_amount = payment_info.installment_amount
                print(f"Next EMI Amount: {next_emi_amount}, Remaining Amount: {remaining_amount}")

                if remaining_amount >= next_emi_amount:
                    # Full payment for the current EMI
                    paid_emi_instance = emi_model(
                        payment_info=payment_info,
                        registration_no=registration_no,
                        date=date,
                        payment_mode=payment_mode,
                        emi=next_emi.emi,
                        amount=next_emi_amount,
                        monthly_payment_type = 'Full payment',
                        status="Paid"
                    )

                    # Set additional fields based on payment mode
                    if payment_mode == 'Bank Transfer':
                        paid_emi_instance.reference_no = request.POST.get('reference_no')
                    elif payment_mode == 'UPI':
                        paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
                        paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

                    paid_emi_instance.save()
                    remaining_amount -= next_emi_amount
                    print(f"Processed Full Payment for {next_emi.emi}: Remaining Amount: {remaining_amount}")

                else:
                    # Partial payment for the current EMI
                    partial_payment_instance = emi_model(
                        payment_info=payment_info,
                        registration_no=registration_no,
                        date=date,
                        payment_mode=payment_mode,
                        emi=next_emi.emi,
                        amount=remaining_amount,
                        monthly_payment_type = 'partial payment',
                        status="Pending"
                    )

                    # Set additional fields based on payment mode
                    if payment_mode == 'Bank Transfer':
                        partial_payment_instance.reference_no = request.POST.get('reference_no')
                    elif payment_mode == 'UPI':
                        partial_payment_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
                        partial_payment_instance.upi_app_name = request.POST.get('upi_app_name')

                    partial_payment_instance.save()
                    next_emi.amount -= remaining_amount
                    remaining_amount = 0  # All remaining amount is paid
                    print(f"Processed Partial Payment for {next_emi.emi}: Remaining Amount: {remaining_amount}")

        messages.success(request, 'Payment processed successfully!')
        return redirect('payment')
    enrollment = Enrollment.objects.all().values()
    context = {
        'next_emi': "EMI_1",
        'enrollment': enrollment,
    }

    return render(request, 'payment_info.html', context)

# def new_installment_update_view(request, id):
#     payment_info = get_object_or_404(PaymentInfo, id=id)
#     print(f"Payment Info: {payment_info}")  # Debug

#     total_fees = payment_info.total_fees
    
#     # Calculate total paid amount by summing over all EMI models
#     total_paid_amount = 0
#     total_pending_amount = 0  # To track pending amount
    
#     # Calculate total paid amount by summing over all EMI models
#     for emi_model in EMI_MODELS.values():
#         # Sum amounts for paid EMIs
#         total_paid_amount += sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info, status='Paid'))

#         # Sum amounts for pending EMIs
#         total_pending_amount += sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info, status='Pending'))

#     remaining_balance = total_fees - (total_paid_amount + total_pending_amount) # Calculate remaining balance

#     print(f"remaining balance : {remaining_balance} , Total Paid Amount : {total_paid_amount} , total_pending_amount : {total_pending_amount}")
    
    
#     if request.method == 'POST':
#         payment_amount = Decimal(request.POST.get('amount', 0))
#         registration_no = request.POST.get('registration_no')
#         payment_mode = request.POST.get('payment_mode') 
#         input_date = request.POST.get('date')
        
#         parsed_date = datetime.strptime(input_date, "%d-%m-%Y")  # Adjust if input format changes
#         date = parsed_date.strftime("%Y-%m-%d")

#         # Validation for payment amount
#         if payment_amount <= 0:
#             messages.error(request, "Invalid payment amount.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         # Ensure payment amount does not exceed remaining balance
#         if payment_amount > remaining_balance:
#             messages.error(request, f"Payment amount exceeds the remaining balance. Remaining balance: {remaining_balance}")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         if not date:
#             messages.error(request, "Date is required.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         try:
#             parsed_date = datetime.strptime(date, '%Y-%m-%d')
#         except ValueError:
#             messages.error(request, "Invalid date format.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         last_pending_emi = get_last_pending_emi(payment_info)
#         emi_type = None
#         remaining_amount = 0

#         if isinstance(last_pending_emi, str):
#             emi_type = last_pending_emi
#         elif hasattr(last_pending_emi, 'emi'):
#             emi_type = last_pending_emi.emi
#             remaining_amount = last_pending_emi.amount

#         if not emi_type:
#             messages.error(request, "No EMIs available for processing.")
#             return redirect('new_installment_update_info', id=payment_info.id)
        
#         installment = payment_info.installment_amount
#         print(f"Installment : {installment}")
        
#         balance = installment
#         print(f"Initial Balance : {balance}")
#         #while start
#         while payment_amount > 0:
#             print(payment_amount)
#             next_emi = get_next_emi(payment_info, emi_type)
            
#             print(f"Next EMI : {next_emi}")
            
#             if next_emi is None:
#                 break

#             emi_model = EMI_MODELS.get(next_emi.emi)
            
#             print(f"EMI Model : {emi_model}")
            
#             if not emi_model:
#                 messages.error(request, f"EMI model not found for {next_emi.emi}.")
#                 return redirect('new_installment_update_info', id=payment_info.id)
            
#             next_emi_amount = sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info))
#             # next_emi_amount = next_emi.amount
#             print(f"Next EMI Amount : {next_emi_amount}")
            
#             balance -= next_emi_amount
#             print(f"After Balance : {balance}")
            
#             if payment_amount >= balance:
#                 paid_emi_instance = emi_model(
#                     payment_info=payment_info,
#                     registration_no=registration_no,
#                     date=parsed_date,
#                     payment_mode=payment_mode,
#                     emi=next_emi.emi,
#                     amount= balance,
#                     status="Paid"
#                 )
#                 paid_emi_instance.save()
#                 payment_amount -= balance
#                 print(f"Payment Amount After Saving : {payment_amount}, Payment paid Amount : {paid_emi_instance.amount}")
                
#                 # next_emi.amount = 0
#             else:
#                 paid_emi_instance = emi_model(
#                     payment_info=payment_info,
#                     registration_no=registration_no,
#                     date=parsed_date,
#                     payment_mode=payment_mode,
#                     emi=next_emi.emi,
#                     amount=payment_amount,
#                     status="Pending"
#                 )
#                 paid_emi_instance.save()

#                 # next_emi.amount -= payment_amount
#                 payment_amount = 0
# #
#         # Add additional fields based on payment mode
#         if payment_mode == 'Bank Transfer':
#             paid_emi_instance.refference_no = request.POST.get('refference_no')
#         elif payment_mode == 'UPI':
#             paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
#             paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

#         try:
#             paid_emi_instance.save()
#             print("Paid EMI instance saved successfully.")  # Debug
#         except Exception as e:
#             print(f"Error saving instance: {e}")  # Debug
#             messages.error(request, "Error saving the payment information.")
#             return redirect('new_installment_update_info', id=payment_info.id)

#         messages.success(request, 'Payment processed successfully.')
#         return redirect('new_manage_payments')

#     # For GET requests or initial form rendering
#     last_pending_emi = get_last_pending_emi(payment_info)

#     if isinstance(last_pending_emi, str):
#         emi_type = last_pending_emi
#     elif hasattr(last_pending_emi, 'emi'):
#         emi_type = last_pending_emi.emi
#     else:
#         messages.error(request, "No EMIs available for processing.")
#         return redirect('new_installment_update_info', id=payment_info.id)

#     # Retrieve installments to display in the table
#     installments = []
#     for emi_model in EMI_MODELS.values():
#         installments.extend(emi_model.objects.filter(payment_info=payment_info))
    
#     context = {
#         'next_emi': emi_type,
#         'payment_info': payment_info,
#         'remaining_balance': remaining_balance,  # Send remaining balance to template
#         'installments': installments  # Assuming this is how you fetch installments
#     }

#     return render(request, 'new_installment_info.html', context)


def new_installment_update_view(request, id):
    payment_info = get_object_or_404(PaymentInfo, id=id)
    print(f"Payment Info: {payment_info}")  # Debug

    total_fees = payment_info.total_fees
    
    # Calculate total paid amount by summing over all EMI models
    total_paid_amount = 0
    total_pending_amount = 0  # To track pending amount
    
    # Calculate total paid amount by summing over all EMI models
    for emi_model in EMI_MODELS.values():
        # Sum amounts for paid EMIs
        total_paid_amount += sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info, status='Paid'))

        # Sum amounts for pending EMIs
        total_pending_amount += sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info, status='Pending'))

    remaining_balance = total_fees - (total_paid_amount + total_pending_amount) # Calculate remaining balance

    print(f"remaining balance : {remaining_balance} , Total Paid Amount : {total_paid_amount} , total_pending_amount : {total_pending_amount}")
    
    
    if request.method == 'POST':
        payment_amount = Decimal(request.POST.get('amount', 0))
        registration_no = request.POST.get('registration_no')
        payment_mode = request.POST.get('payment_mode') 
        input_date = request.POST.get('date')
        
        parsed_date = datetime.strptime(input_date, "%d-%m-%Y")  # Adjust if input format changes
        date = parsed_date.strftime("%Y-%m-%d")

        # Validation for payment amount
        if payment_amount <= 0:
            messages.error(request, "Invalid payment amount.")
            return redirect('new_installment_update_info', id=payment_info.id)

        # Ensure payment amount does not exceed remaining balance
        if payment_amount > remaining_balance:
            messages.error(request, f"Payment amount exceeds the remaining balance. Remaining balance: {remaining_balance}")
            return redirect('new_installment_update_info', id=payment_info.id)

        if not date:
            messages.error(request, "Date is required.")
            return redirect('new_installment_update_info', id=payment_info.id)

        try:
            parsed_date = datetime.strptime(date, '%Y-%m-%d')
        except ValueError:
            messages.error(request, "Invalid date format.")
            return redirect('new_installment_update_info', id=payment_info.id)

        last_pending_emi = get_last_pending_emi(payment_info)
        emi_type = None
        remaining_amount = 0

        if isinstance(last_pending_emi, str):
            emi_type = last_pending_emi
        elif hasattr(last_pending_emi, 'emi'):
            emi_type = last_pending_emi.emi
            remaining_amount = last_pending_emi.amount

        if not emi_type:
            messages.error(request, "No EMIs available for processing.")
            return redirect('new_installment_update_info', id=payment_info.id)
        
        installment = payment_info.installment_amount
        print(f"Installment : {installment}")
        
        balance = installment
        print(f"Initial Balance : {balance}")
        #while start
        while payment_amount > 0:
            print(payment_amount)
            next_emi = get_next_emi(payment_info, emi_type)
            
            print(f"Next EMI : {next_emi}")
            
            if next_emi is None:
                break

            emi_model = EMI_MODELS.get(next_emi.emi)
            
            print(f"EMI Model : {emi_model}")
            
            if not emi_model:
                messages.error(request, f"EMI model not found for {next_emi.emi}.")
                return redirect('new_installment_update_info', id=payment_info.id)
            
            next_emi_amount = sum(emi.amount for emi in emi_model.objects.filter(payment_info=payment_info))
            # next_emi_amount = next_emi.amount
            print(f"Next EMI Amount : {next_emi_amount}")
            
            balance = abs(balance - next_emi_amount)
            print(f"After Balance : {balance}")
            
            if payment_amount < balance:
                paid_emi_instance = emi_model(
                    payment_info=payment_info,
                    registration_no=registration_no,
                    date=parsed_date,
                    payment_mode=payment_mode,
                    emi=next_emi.emi,
                    amount=payment_amount,
                    status="Pending"
                )
                paid_emi_instance.save()
                payment_amount = 0
                # next_emi.amount = 0
            else:
                paid_emi_instance = emi_model(
                    payment_info=payment_info,
                    registration_no=registration_no,
                    date=parsed_date,
                    payment_mode=payment_mode,
                    emi=next_emi.emi,
                    amount= balance,
                    status="Paid"
                )
                paid_emi_instance.save()
                payment_amount -= balance
                print(f"Payment Amount After Saving : {payment_amount}, Payment paid Amount : {paid_emi_instance.amount}")
                
                

                # next_emi.amount -= payment_amount
#
        # Add additional fields based on payment mode
        if payment_mode == 'Bank Transfer':
            paid_emi_instance.refference_no = request.POST.get('refference_no')
        elif payment_mode == 'UPI':
            paid_emi_instance.upi_transaction_id = request.POST.get('upi_transaction_id')
            paid_emi_instance.upi_app_name = request.POST.get('upi_app_name')

        try:
            paid_emi_instance.save()
            print("Paid EMI instance saved successfully.")  # Debug
        except Exception as e:
            print(f"Error saving instance: {e}")  # Debug
            messages.error(request, "Error saving the payment information.")
            return redirect('new_installment_update_info', id=payment_info.id)

        messages.success(request, 'Payment processed successfully.')
        return redirect('new_manage_payments')

    # For GET requests or initial form rendering
    last_pending_emi = get_last_pending_emi(payment_info)

    if isinstance(last_pending_emi, str):
        emi_type = last_pending_emi
    elif hasattr(last_pending_emi, 'emi'):
        emi_type = last_pending_emi.emi
    else:
        messages.error(request, "No EMIs available for processing.")
        return redirect('new_installment_update_info', id=payment_info.id)

    # Retrieve installments to display in the table
    installments = []
    for emi_model in EMI_MODELS.values():
        installments.extend(emi_model.objects.filter(payment_info=payment_info))
    
    context = {
        'next_emi': emi_type,
        'payment_info': payment_info,
        'remaining_balance': remaining_balance,  # Send remaining balance to template
        'installments': installments  # Assuming this is how you fetch installments
    }

    return render(request, 'new_installment_info.html', context)
